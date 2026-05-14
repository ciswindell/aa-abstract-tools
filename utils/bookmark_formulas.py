#!/usr/bin/env python3
"""
Bookmark formula utilities for Excel templates.
"""

import pandas as pd
from openpyxl import load_workbook


def _index_to_col_letter(position: int) -> str:
    """Convert 0-based column index to Excel column letter (A, B, ..., AA, AB, ...)."""
    if position < 0:
        return ""

    result = ""
    col = position
    while col >= 0:
        result = chr(col % 26 + ord("A")) + result
        col = col // 26 - 1
    return result


def _header_positions(ws) -> dict[str, int]:
    """Return lowercase header -> 1-based column index mapping for first row."""
    return {
        str(cell.value).strip().lower(): idx + 1
        for idx, cell in enumerate(ws[1])
        if cell.value
    }


def apply_bookmark_formulas(
    workbook_path: str, df: pd.DataFrame, bookmark_column: str
) -> None:
    """Apply bookmark formulas to a specific column in an Excel workbook.

    Creates formulas like: =A2&"-"&B2&"-"&TEXT(C2,"m/d/yyyy")
    Combining Index#, Document Type, and Received Date.
    """
    try:
        wb = load_workbook(workbook_path)
        ws = wb.active

        # Find column positions by header names (case-insensitive)
        headers = _header_positions(ws)

        bookmark_col = headers.get(bookmark_column.strip().lower())
        index_col = headers.get("index#")
        doc_type_col = headers.get("document type")
        received_date_col = headers.get("received date")

        if not all([bookmark_col, index_col, doc_type_col, received_date_col]):
            return  # Missing required columns

        # Generate formulas for each data row
        for row_num in range(2, len(df) + 2):
            formula = (
                f'={_index_to_col_letter(index_col - 1)}{row_num}&"-"&'
                f'{_index_to_col_letter(doc_type_col - 1)}{row_num}&"-"&'
                f'TEXT({_index_to_col_letter(received_date_col - 1)}{row_num},"m/d/yyyy")'
            )
            ws.cell(row=row_num, column=bookmark_col, value=formula)

        wb.save(workbook_path)
        wb.close()

    except Exception:
        pass  # Skip if formulas can't be applied


def has_bookmark_formulas(workbook_path: str, bookmark_column: str) -> bool:
    """Check if bookmark column already has formulas."""
    try:
        wb = load_workbook(workbook_path)
        ws = wb.active

        # Find bookmark column
        headers = _header_positions(ws)
        bookmark_col = headers.get(bookmark_column.strip().lower())

        if bookmark_col:
            # Check if row 2 has a formula
            cell = ws.cell(row=2, column=bookmark_col)
            has_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            wb.close()
            return has_formula

        wb.close()
        return False

    except Exception:
        return False


def detect_bookmark_column(df: pd.DataFrame) -> str | None:
    """Detect bookmark formula column in DataFrame."""
    candidates = ["Bookmark Formula", "Bookmark", "Bookmark Text", "Formula"]
    for col in df.columns:
        for candidate in candidates:
            if str(col).strip().lower() == candidate.lower():
                return str(col)
    return None
