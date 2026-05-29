#!/usr/bin/env python3
"""End-to-end tests for the Orphan marker column (#2)."""

import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter

from adapters.excel_repo import ExcelOpenpyxlRepo
from adapters.pdf_repo import PypdfPdfRepo
from core.pipeline.context import PipelineContext
from core.pipeline.steps.load_step import LoadStep
from core.pipeline.steps.save_step import SaveStep
from core.pipeline.steps.validate_step import ValidateStep


class _Logger:
    def info(self, *a, **k):
        pass

    def warning(self, *a, **k):
        pass

    def error(self, *a, **k):
        pass


class _YesUI:
    def prompt_add_new_bookmarks(self, titles):
        return True


def _headers(ws):
    return [c.value for c in ws[1] if c.value not in (None, "")]


def _col_index(ws, name):
    for i, c in enumerate(ws[1], start=1):
        if c.value == name:
            return i
    raise AssertionError(f"column {name!r} not found in {_headers(ws)}")


def test_orphan_column_written_for_added_rows():
    excel_repo, pdf_repo, logger = ExcelOpenpyxlRepo(), PypdfPdfRepo(), _Logger()
    with tempfile.TemporaryDirectory() as d:
        dpath = Path(d)
        xlsx = dpath / "in.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        # Real templates already include Document_Found, so Orphan is the only
        # new column appended -> it lands at the end.
        ws.append(
            [
                "Index#",
                "Document Type",
                "Received Date",
                "Legal Description",
                "Grantee",
                "Grantor",
                "Document Date",
                "Document_Found",
            ]
        )
        ws.append([1, "Deed", "2020-01-01", "L1", "A", "B", "2020-01-01", "Yes"])
        ws.append([2, "Release", "2020-02-02", "L2", "C", "D", "2020-02-02", "Yes"])
        wb.save(xlsx)

        pdf = dpath / "in.pdf"
        w = PdfWriter()
        for _ in range(3):
            w.add_blank_page(width=200, height=200)
        w.add_outline_item("1-Deed-1/1/2020", 0)
        w.add_outline_item("2-Release-2/2/2020", 1)
        w.add_outline_item("Merger 2014", 2)  # orphan
        with open(pdf, "wb") as fh:
            w.write(fh)

        ctx = PipelineContext(
            file_pairs=[(str(xlsx), str(pdf), "Sheet")],
            options={"sheet_name": "Sheet", "backup": False},
        )
        ValidateStep(excel_repo, pdf_repo, logger, _YesUI()).execute(ctx)
        LoadStep(excel_repo, pdf_repo, logger, None).execute(ctx)

        out_xlsx = dpath / "out.xlsx"
        ctx.excel_out_path = str(out_xlsx)
        ctx.pdf_out_path = str(dpath / "out.pdf")
        ctx.final_pdf = PdfWriter()
        SaveStep(excel_repo, pdf_repo, logger, None).execute(ctx)

        wb2 = load_workbook(str(out_xlsx))
        ws2 = wb2.active
        headers = _headers(ws2)
        orphan_i = _col_index(ws2, "Orphan")
        dt_i = _col_index(ws2, "Document Type")
        by_dt = {}
        for r in range(2, ws2.max_row + 1):
            dt = ws2.cell(row=r, column=dt_i).value
            if dt in (None, ""):
                continue
            by_dt[dt] = ws2.cell(row=r, column=orphan_i).value
        wb2.close()

        assert "Orphan" in headers
        assert headers[-1] == "Orphan"  # appended at the end
        assert by_dt["Merger 2014"] == "Yes"
        assert by_dt["Deed"] == "No"
        assert by_dt["Release"] == "No"


def test_existing_orphan_column_untouched_when_no_orphans():
    excel_repo, pdf_repo, logger = ExcelOpenpyxlRepo(), PypdfPdfRepo(), _Logger()
    with tempfile.TemporaryDirectory() as d:
        dpath = Path(d)
        xlsx = dpath / "in.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        # Input already carries an Orphan column from a prior run.
        ws.append(
            [
                "Index#",
                "Document Type",
                "Received Date",
                "Legal Description",
                "Grantee",
                "Grantor",
                "Document Date",
                "Orphan",
            ]
        )
        ws.append([1, "Deed", "2020-01-01", "L1", "A", "B", "2020-01-01", "Yes"])
        ws.append([2, "Release", "2020-02-02", "L2", "C", "D", "2020-02-02", "No"])
        wb.save(xlsx)

        pdf = dpath / "in.pdf"
        w = PdfWriter()
        for _ in range(2):
            w.add_blank_page(width=200, height=200)
        w.add_outline_item("1-Deed-1/1/2020", 0)
        w.add_outline_item("2-Release-2/2/2020", 1)  # all match -> no orphans
        with open(pdf, "wb") as fh:
            w.write(fh)

        ctx = PipelineContext(
            file_pairs=[(str(xlsx), str(pdf), "Sheet")],
            options={"sheet_name": "Sheet", "backup": False},
        )
        # No orphans -> prompt never fires; titles stay None.
        ValidateStep(excel_repo, pdf_repo, logger, _YesUI()).execute(ctx)
        assert ctx.new_bookmark_titles is None
        LoadStep(excel_repo, pdf_repo, logger, None).execute(ctx)

        out_xlsx = dpath / "out.xlsx"
        ctx.excel_out_path = str(out_xlsx)
        ctx.pdf_out_path = str(dpath / "out.pdf")
        ctx.final_pdf = PdfWriter()
        SaveStep(excel_repo, pdf_repo, logger, None).execute(ctx)

        wb2 = load_workbook(str(out_xlsx))
        ws2 = wb2.active
        orphan_i = _col_index(ws2, "Orphan")
        dt_i = _col_index(ws2, "Document Type")
        by_dt = {}
        for r in range(2, ws2.max_row + 1):
            dt = ws2.cell(row=r, column=dt_i).value
            if dt in (None, ""):
                continue
            by_dt[dt] = ws2.cell(row=r, column=orphan_i).value
        wb2.close()

        # The pre-existing Orphan values are round-tripped unchanged.
        assert by_dt["Deed"] == "Yes"
        assert by_dt["Release"] == "No"


def test_existing_orphan_column_overwritten_in_place_no_duplicate():
    excel_repo, pdf_repo, logger = ExcelOpenpyxlRepo(), PypdfPdfRepo(), _Logger()
    with tempfile.TemporaryDirectory() as d:
        dpath = Path(d)
        xlsx = dpath / "in.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        # Input already has an Orphan column from a prior run (Deed marked Yes).
        ws.append(
            [
                "Index#",
                "Document Type",
                "Received Date",
                "Legal Description",
                "Grantee",
                "Grantor",
                "Document Date",
                "Orphan",
            ]
        )
        ws.append([1, "Deed", "2020-01-01", "L1", "A", "B", "2020-01-01", "Yes"])
        ws.append([2, "Release", "2020-02-02", "L2", "C", "D", "2020-02-02", "No"])
        wb.save(xlsx)

        pdf = dpath / "in.pdf"
        w = PdfWriter()
        for _ in range(3):
            w.add_blank_page(width=200, height=200)
        w.add_outline_item("1-Deed-1/1/2020", 0)
        w.add_outline_item("2-Release-2/2/2020", 1)
        w.add_outline_item("Court Order", 2)  # NEW orphan this run
        with open(pdf, "wb") as fh:
            w.write(fh)

        ctx = PipelineContext(
            file_pairs=[(str(xlsx), str(pdf), "Sheet")],
            options={"sheet_name": "Sheet", "backup": False},
        )
        ValidateStep(excel_repo, pdf_repo, logger, _YesUI()).execute(ctx)
        assert ctx.new_bookmark_titles == {"Court Order"}
        LoadStep(excel_repo, pdf_repo, logger, None).execute(ctx)

        out_xlsx = dpath / "out.xlsx"
        ctx.excel_out_path = str(out_xlsx)
        ctx.pdf_out_path = str(dpath / "out.pdf")
        ctx.final_pdf = PdfWriter()
        SaveStep(excel_repo, pdf_repo, logger, None).execute(ctx)

        wb2 = load_workbook(str(out_xlsx))
        ws2 = wb2.active
        headers = _headers(ws2)
        orphan_i = _col_index(ws2, "Orphan")
        dt_i = _col_index(ws2, "Document Type")
        by_dt = {}
        for r in range(2, ws2.max_row + 1):
            dt = ws2.cell(row=r, column=dt_i).value
            if dt in (None, ""):
                continue
            by_dt[dt] = ws2.cell(row=r, column=orphan_i).value
        wb2.close()

        # Exactly one Orphan column — no duplicate created on the second run.
        assert headers.count("Orphan") == 1
        # Overwrite reflects THIS run: prior "Yes" on Deed is reset to "No";
        # only this run's new orphan is "Yes".
        assert by_dt["Deed"] == "No"
        assert by_dt["Release"] == "No"
        assert by_dt["Court Order"] == "Yes"
