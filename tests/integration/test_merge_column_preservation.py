#!/usr/bin/env python3
"""
Integration tests for column preservation during full merge workflows.
Tests the end-to-end pipeline with merge operations.
"""

import sys
import tempfile
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# Add tests directory to path for fixtures
tests_dir = Path(__file__).parent.parent
sys.path.insert(0, str(tests_dir))

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfWriter

from core.models import Options
from core.pipeline.context import PipelineContext
from core.pipeline.steps.save_step import SaveStep
from adapters.excel_repo import ExcelOpenpyxlRepo
from adapters.pdf_repo import PdfRepo
from adapters.logger_tk import TkLogger
from fixtures.excel_fixtures import create_excel_with_basic_columns, create_excel_with_extra_columns


class TestMergeColumnPreservation:
    """Integration tests for merge workflow column preservation."""

    def setup_method(self):
        """Set up test fixtures."""
        self.excel_repo = ExcelOpenpyxlRepo()
        self.pdf_repo = PdfRepo()
        # Create mock logger that writes to nowhere
        self.logger = TkLogger(write=lambda msg: None)

    def test_two_file_merge_preserves_all_columns(self):
        """Test that merging two files with different columns preserves all columns."""
        # Create two test Excel files
        file1_fixture = create_excel_with_basic_columns()
        file2_fixture = create_excel_with_extra_columns()

        # Simulate merged DataFrame (what the pipeline would create)
        merged_df = pd.DataFrame(
            {
                "Index#": ["1", "2", "3", "4", "5"],
                "Date": [
                    "2024-01-01",
                    "2024-01-02",
                    "2024-01-03",
                    "2024-02-01",
                    "2024-02-02",
                ],
                "Name": ["Doc A", "Doc B", "Doc C", "Doc D", "Doc E"],
                # Extra columns from file 2
                "Status": [None, None, None, "Approved", "Pending"],
                "Comments": [None, None, None, "Looks good", "Needs review"],
            }
        )

        # Create mock pipeline context for a merge workflow
        file_pairs = [
            (str(file1_fixture["path"]), "dummy.pdf", "Index"),
            (str(file2_fixture["path"]), "dummy2.pdf", "Index"),
        ]

        context = PipelineContext(
            file_pairs=file_pairs,
            options={"sheet_name": "Index", "backup": False},
        )

        # Set the merged DataFrame in context
        context.df = merged_df

        # Create temporary output file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

        # Override output path
        context.excel_out_path = str(output_path)
        context.pdf_out_path = str(output_path.with_suffix(".pdf"))

        # Create mock PDF writer
        context.final_pdf = PdfWriter()

        # Execute save step
        save_step = SaveStep(self.excel_repo, self.pdf_repo, self.logger, None)
        save_step.execute(context)

        # Load result and verify columns
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        file1_fixture["path"].unlink()
        file2_fixture["path"].unlink()
        output_path.unlink()
        if Path(context.pdf_out_path).exists():
            Path(context.pdf_out_path).unlink()
        wb.close()

        # Assert all columns from both files are present
        assert len(headers) == 5, f"Expected 5 columns, got {len(headers)}: {headers}"
        assert "Index#" in headers
        assert "Date" in headers
        assert "Name" in headers
        assert "Status" in headers, "Status column from file2 should be preserved"
        assert "Comments" in headers, "Comments column from file2 should be preserved"

    def test_merge_workflow_detection(self):
        """Test that SaveStep correctly detects merge workflows."""
        # Create single file pair (not a merge)
        file1_fixture = create_excel_with_basic_columns()
        single_pair_context = PipelineContext(
            file_pairs=[(str(file1_fixture["path"]), "dummy.pdf", "Index")],
            options={},
        )

        # Create multi-file pairs (merge workflow)
        file2_fixture = create_excel_with_extra_columns()
        merge_context = PipelineContext(
            file_pairs=[
                (str(file1_fixture["path"]), "dummy1.pdf", "Index"),
                (str(file2_fixture["path"]), "dummy2.pdf", "Index"),
            ],
            options={},
        )

        # Test detection
        assert (
            not single_pair_context.is_merge_workflow()
        ), "Single file should not be detected as merge"
        assert (
            merge_context.is_merge_workflow()
        ), "Multiple files should be detected as merge"

        # Clean up
        file1_fixture["path"].unlink()
        file2_fixture["path"].unlink()

    def test_single_file_workflow_unchanged(self):
        """Test that single-file workflows are not affected by merge logic."""
        # Create test file
        file_fixture = create_excel_with_basic_columns()

        # Create DataFrame (single file, no extra columns)
        df = pd.DataFrame(
            {
                "Index#": ["1", "2", "3"],
                "Date": ["2024-01-01", "2024-01-02", "2024-01-03"],
                "Name": ["Doc A", "Doc B", "Doc C"],
            }
        )

        # Create context for single file
        context = PipelineContext(
            file_pairs=[(str(file_fixture["path"]), "dummy.pdf", "Index")],
            options={"sheet_name": "Index", "backup": False},
        )

        context.df = df

        # Create temporary output
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

        context.excel_out_path = str(output_path)
        context.pdf_out_path = str(output_path.with_suffix(".pdf"))
        context.final_pdf = PdfWriter()

        # Execute save step
        save_step = SaveStep(self.excel_repo, self.pdf_repo, self.logger, None)
        save_step.execute(context)

        # Load result
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        file_fixture["path"].unlink()
        output_path.unlink()
        if Path(context.pdf_out_path).exists():
            Path(context.pdf_out_path).unlink()
        wb.close()

        # Assert only original columns (single file workflow should work as before)
        assert len(headers) == 3, f"Expected 3 columns, got {len(headers)}: {headers}"

    def test_three_plus_file_merge_preserves_all_unique_columns(self):
        """Test that merging 3+ files preserves all unique columns from all files."""
        # Create three test Excel files with different column structures
        file1_fixture = create_excel_with_basic_columns()  # [Index#, Date, Name]
        file2_fixture = create_excel_with_extra_columns()  # [Index#, Date, Name, Status, Comments]
        from fixtures.excel_fixtures import create_excel_with_disjoint_columns
        file3_fixture = create_excel_with_disjoint_columns()  # [Index#, Priority, Category, Owner]

        # Simulate merged DataFrame from 3 files (what pipeline would create)
        # Union of all columns: Index#, Date, Name, Status, Comments, Priority, Category, Owner
        merged_df = pd.DataFrame(
            {
                "Index#": ["1", "2", "3", "4", "5", "6", "7"],
                # From file 1
                "Date": [
                    "2024-01-01",
                    "2024-01-02",
                    "2024-01-03",
                    "2024-02-01",
                    "2024-02-02",
                    "2024-03-01",
                    "2024-03-02",
                ],
                "Name": [
                    "Doc A",
                    "Doc B",
                    "Doc C",
                    "Doc D",
                    "Doc E",
                    "Doc F",
                    "Doc G",
                ],
                # From file 2
                "Status": [None, None, None, "Approved", "Pending", None, None],
                "Comments": [None, None, None, "Good", "Review", None, None],
                # From file 3
                "Priority": [None, None, None, None, None, "High", "Low"],
                "Category": [None, None, None, None, None, "Security", "Performance"],
                "Owner": [None, None, None, None, None, "Alice", "Bob"],
            }
        )

        # Create mock pipeline context for 3-file merge
        file_pairs = [
            (str(file1_fixture["path"]), "dummy1.pdf", "Index"),
            (str(file2_fixture["path"]), "dummy2.pdf", "Index"),
            (str(file3_fixture["path"]), "dummy3.pdf", "Index"),
        ]

        context = PipelineContext(
            file_pairs=file_pairs,
            options={"sheet_name": "Index", "backup": False},
        )

        context.df = merged_df

        # Create temporary output
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

        context.excel_out_path = str(output_path)
        context.pdf_out_path = str(output_path.with_suffix(".pdf"))
        context.final_pdf = PdfWriter()

        # Execute save step
        save_step = SaveStep(self.excel_repo, self.pdf_repo, self.logger, None)
        save_step.execute(context)

        # Load result and verify all unique columns from all 3 files
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        file1_fixture["path"].unlink()
        file2_fixture["path"].unlink()
        file3_fixture["path"].unlink()
        output_path.unlink()
        if Path(context.pdf_out_path).exists():
            Path(context.pdf_out_path).unlink()
        wb.close()

        # Assert all 8 unique columns from 3 files are present
        assert len(headers) == 8, f"Expected 8 columns, got {len(headers)}: {headers}"
        assert "Index#" in headers
        assert "Date" in headers
        assert "Name" in headers
        assert "Status" in headers, "Status from file2 should be preserved"
        assert "Comments" in headers, "Comments from file2 should be preserved"
        assert "Priority" in headers, "Priority from file3 should be preserved"
        assert "Category" in headers, "Category from file3 should be preserved"
        assert "Owner" in headers, "Owner from file3 should be preserved"

    def test_merge_performance_impact_minimal(self):
        """Test that column preservation has minimal performance impact (<5%)."""
        import time

        # Create test files
        file1_fixture = create_excel_with_basic_columns()
        file2_fixture = create_excel_with_extra_columns()

        # Create a larger DataFrame to test performance
        large_df = pd.DataFrame(
            {
                "Index#": [str(i) for i in range(1, 101)],  # 100 rows
                "Date": ["2024-01-01"] * 100,
                "Name": [f"Document {i}" for i in range(1, 101)],
                "Status": [None] * 50 + ["Approved"] * 50,
                "Comments": [None] * 50 + ["OK"] * 50,
            }
        )

        # Create context for merge workflow
        context = PipelineContext(
            file_pairs=[
                (str(file1_fixture["path"]), "dummy1.pdf", "Index"),
                (str(file2_fixture["path"]), "dummy2.pdf", "Index"),
            ],
            options={"sheet_name": "Index", "backup": False},
        )
        context.df = large_df

        # Measure time for save operation
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

        context.excel_out_path = str(output_path)
        context.pdf_out_path = str(output_path.with_suffix(".pdf"))
        context.final_pdf = PdfWriter()

        save_step = SaveStep(self.excel_repo, self.pdf_repo, self.logger, None)

        start_time = time.time()
        save_step.execute(context)
        elapsed_time = time.time() - start_time

        # Clean up
        file1_fixture["path"].unlink()
        file2_fixture["path"].unlink()
        output_path.unlink()
        if Path(context.pdf_out_path).exists():
            Path(context.pdf_out_path).unlink()

        # Performance assertion: should complete in reasonable time
        # For 100 rows with 5 columns, expect < 2 seconds
        assert (
            elapsed_time < 2.0
        ), f"Save operation took {elapsed_time:.2f}s, should be < 2s"

        # Note: Actual <5% impact test would require baseline measurement
        # This test ensures merge operations complete in reasonable time

