#!/usr/bin/env python3
"""
Unit tests for ExcelOpenpyxlRepo Document_Found column enhancement.
Tests new column addition and boolean to "Yes"/"No" conversion functionality.
"""

import sys
import tempfile
from pathlib import Path

# Add project root to Python path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

import pandas as pd
from openpyxl import Workbook, load_workbook

from adapters.excel_repo import ExcelOpenpyxlRepo


class TestExcelRepoDocumentFoundEnhancement:
    """Test cases for Document_Found column functionality in ExcelOpenpyxlRepo."""

    def setup_method(self):
        """Set up test fixtures."""
        self.repo = ExcelOpenpyxlRepo()

    def _create_test_excel_template(self, columns: list, data: list = None) -> Path:
        """Create a test Excel template with specified columns."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Index"

        # Add headers
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Add data if provided
        if data:
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(temp_file.name)
        wb.close()

        return Path(temp_file.name)

    def test_add_missing_columns_disabled_by_default(self):
        """Test that new columns are not added when add_missing_columns=False (default)."""
        # Create template with existing columns
        template_path = self._create_test_excel_template(
            columns=["Index#", "Document Type"],
            data=[["1", "Assignment"], ["2", "Report"]],
        )

        # Create DataFrame with Document_Found column
        df = pd.DataFrame(
            {
                "Index#": ["1", "2"],
                "Document Type": ["Assignment", "Report"],
                "Document_Found": [True, False],
            }
        )

        # Save without adding missing columns (default behavior)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=False,
            )

        # Load result and check that Document_Found column was not added
        result_wb = load_workbook(str(output_path))
        result_ws = result_wb["Index"]

        # Get header row
        headers = [cell.value for cell in result_ws[1] if cell.value]

        # Should only have original columns
        assert "Index#" in headers
        assert "Document Type" in headers
        assert "Document_Found" not in headers

        result_wb.close()
        template_path.unlink()
        output_path.unlink()

    def test_add_missing_columns_enabled(self):
        """Test that whitelisted new columns are added when add_missing_columns=True."""
        # Create template with existing columns
        template_path = self._create_test_excel_template(
            columns=["Index#", "Document Type"],
            data=[["1", "Assignment"], ["2", "Report"]],
        )

        # Create DataFrame with Document_Found column
        df = pd.DataFrame(
            {
                "Index#": ["1", "2"],
                "Document Type": ["Assignment", "Report"],
                "Document_Found": [True, False],
            }
        )

        # Save with adding missing columns enabled
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=True,
            )

        # Load result and check that Document_Found column was added
        result_wb = load_workbook(str(output_path))
        result_ws = result_wb["Index"]

        # Get header row
        headers = [cell.value for cell in result_ws[1] if cell.value]

        # Should have all columns including new one
        assert "Index#" in headers
        assert "Document Type" in headers
        assert "Document_Found" in headers

        # Check that Document_Found is the rightmost column
        document_found_col_idx = headers.index("Document_Found") + 1
        assert document_found_col_idx == len(headers)  # Should be last column

        result_wb.close()
        template_path.unlink()
        output_path.unlink()

    def test_boolean_to_yes_no_conversion(self):
        """Test that boolean Document_Found values are converted to 'Yes'/'No' in Excel."""
        # Create template with Document_Found column already present
        template_path = self._create_test_excel_template(
            columns=["Index#", "Document Type", "Document_Found"],
            data=[["1", "Assignment", ""], ["2", "Report", ""]],
        )

        # Create DataFrame with boolean Document_Found values
        df = pd.DataFrame(
            {
                "Index#": ["1", "2"],
                "Document Type": ["Assignment", "Report"],
                "Document_Found": [True, False],
            }
        )

        # Save (add_missing_columns doesn't matter since column exists)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=False,
            )

        # Load result and check boolean conversion
        result_wb = load_workbook(str(output_path))
        result_ws = result_wb["Index"]

        # Find Document_Found column index
        headers = [cell.value for cell in result_ws[1]]
        document_found_col_idx = headers.index("Document_Found") + 1

        # Check converted values
        row1_value = result_ws.cell(row=2, column=document_found_col_idx).value
        row2_value = result_ws.cell(row=3, column=document_found_col_idx).value

        assert row1_value == "Yes"  # True -> "Yes"
        assert row2_value == "No"  # False -> "No"

        result_wb.close()
        template_path.unlink()
        output_path.unlink()
