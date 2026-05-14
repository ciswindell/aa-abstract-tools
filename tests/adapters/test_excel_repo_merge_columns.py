#!/usr/bin/env python3
"""
Unit tests for ExcelRepo column preservation during merge operations.
Tests column preservation, system column exclusion, and whitespace normalization.
"""

import tempfile
from pathlib import Path

import pandas as pd
from fixtures.excel_fixtures import (
    create_excel_with_basic_columns,
)
from openpyxl import load_workbook

from adapters.excel_repo import ExcelOpenpyxlRepo


class TestExcelRepoColumnPreservation:
    """Test column preservation during merge operations."""

    def setup_method(self):
        """Set up test fixtures."""
        self.repo = ExcelOpenpyxlRepo()

    def test_preserve_extra_columns_from_merge(self):
        """Test that new columns are added when add_missing_columns=True."""
        # Create template with basic columns [Index#, Date, Name]
        template_fixture = create_excel_with_basic_columns()
        template_path = template_fixture["path"]

        # Create DataFrame with additional columns (simulating merged data)
        df = pd.DataFrame(
            {
                "Index#": ["1", "2", "3", "4", "5"],
                "Date": [
                    "2024-01-01",
                    "2024-01-02",
                    "2024-01-03",
                    "2024-02-01",
                    "2024-02-02",
                ],
                "Name": ["Doc A", "Doc B", "Doc C", "Doc D", "Doc E"],
                "Status": ["Approved", "Approved", "Approved", "Pending", "Pending"],
                "Comments": ["OK", "OK", "OK", "Review", "Review"],
            }
        )

        # Save with add_missing_columns=True (merge workflow behavior)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=True,
            )

        # Load result and verify all columns are present
        wb = load_workbook(str(output_path))
        ws = wb.active

        # Get headers from row 1
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert all 5 columns are present
        assert len(headers) == 5, f"Expected 5 columns, got {len(headers)}: {headers}"
        assert "Index#" in headers
        assert "Date" in headers
        assert "Name" in headers
        assert "Status" in headers, "Status column should be added"
        assert "Comments" in headers, "Comments column should be added"

    def test_preserve_extra_columns_disabled_by_default(self):
        """Test that new columns are NOT added when add_missing_columns=False."""
        # Create template with basic columns
        template_fixture = create_excel_with_basic_columns()
        template_path = template_fixture["path"]

        # Create DataFrame with additional columns
        df = pd.DataFrame(
            {
                "Index#": ["1", "2"],
                "Date": ["2024-01-01", "2024-01-02"],
                "Name": ["Doc A", "Doc B"],
                "Status": ["Approved", "Pending"],
            }
        )

        # Save with add_missing_columns=False (single-file workflow behavior)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=False,
            )

        # Load result
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert only template columns are present
        assert len(headers) == 3, f"Expected 3 columns, got {len(headers)}: {headers}"
        assert "Status" not in headers, "Status should not be added"

    def test_exclude_system_columns(self):
        """Test that system columns are excluded from output."""
        # Create template with basic columns
        template_fixture = create_excel_with_basic_columns()
        template_path = template_fixture["path"]

        # Create DataFrame with system columns
        df = pd.DataFrame(
            {
                "Index#": ["1", "2"],
                "Date": ["2024-01-01", "2024-01-02"],
                "Name": ["Doc A", "Doc B"],
                "_include": [True, False],  # System column
                "Document_ID": ["file1_001", "file1_002"],  # System column
                "_original_index": [0, 1],  # System column
                "Status": ["Approved", "Pending"],  # User column
            }
        )

        # Save with add_missing_columns=True
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=True,
            )

        # Load result
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert system columns are excluded, user column is included
        assert "_include" not in headers, "System column _include should be excluded"
        assert "Document_ID" not in headers, (
            "System column Document_ID should be excluded"
        )
        assert "_original_index" not in headers, (
            "System column _original_index should be excluded"
        )
        assert "Status" in headers, "User column Status should be included"
        assert len(headers) == 4, f"Expected 4 columns, got {len(headers)}: {headers}"

    def test_case_insensitive_column_matching(self):
        """Test that column matching is case-insensitive."""
        # Create template with mixed case columns
        template_fixture = create_excel_with_basic_columns()
        template_path = template_fixture["path"]

        # Create DataFrame with different case
        df = pd.DataFrame(
            {
                "index#": ["1", "2"],  # lowercase
                "DATE": ["2024-01-01", "2024-01-02"],  # uppercase
                "Name": ["Doc A", "Doc B"],  # mixed case
            }
        )

        # Save
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=False,
            )

        # Load result and verify data was written (case-insensitive match worked)
        wb = load_workbook(str(output_path))
        ws = wb.active

        # Check that data was written to row 2
        index_val = ws.cell(row=2, column=1).value
        date_val = ws.cell(row=2, column=2).value

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert data was written (columns matched despite case differences)
        assert index_val == "1", "Data should be written despite case differences"
        assert date_val is not None, "Date should be written despite case differences"

    def test_whitespace_normalization_in_column_names(self):
        """Test that column names with varying whitespace are matched correctly."""
        # Create template with normal column names
        template_fixture = create_excel_with_basic_columns()
        template_path = template_fixture["path"]

        # Create DataFrame with whitespace variations
        df = pd.DataFrame(
            {
                " Index# ": ["1", "2"],  # leading/trailing spaces
                "Date  ": ["2024-01-01", "2024-01-02"],  # trailing spaces
                "  Name": ["Doc A", "Doc B"],  # leading spaces
            }
        )

        # Save
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=False,
            )

        # Load result and verify data was written
        wb = load_workbook(str(output_path))
        ws = wb.active
        index_val = ws.cell(row=2, column=1).value

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert data was written (whitespace normalized)
        assert index_val == "1", "Data should be written despite whitespace differences"

    def test_multi_file_column_union(self):
        """Test that merging multiple files creates a union of all columns."""
        # Create template with columns [Index#, Date, Name]
        template_fixture = create_excel_with_basic_columns()
        template_path = template_fixture["path"]

        # Simulate merged DataFrame from 3 files with different columns
        # File 1: [Index#, Date, Name]
        # File 2: [Index#, Date, Name, Status]
        # File 3: [Index#, Date, Name, Category, Priority]
        df = pd.DataFrame(
            {
                "Index#": ["1", "2", "3", "4", "5"],
                "Date": [
                    "2024-01-01",
                    "2024-01-02",
                    "2024-01-03",
                    "2024-02-01",
                    "2024-02-02",
                ],
                "Name": ["Doc A", "Doc B", "Doc C", "Doc D", "Doc E"],
                "Status": [None, None, None, "Approved", None],  # From file 2
                "Category": [None, None, None, None, "Security"],  # From file 3
                "Priority": [None, None, None, None, "High"],  # From file 3
            }
        )

        # Save with add_missing_columns=True
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=True,
            )

        # Load result and verify all columns from all 3 files are present
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert all 6 unique columns are present (union of all files)
        assert len(headers) == 6, f"Expected 6 columns, got {len(headers)}: {headers}"
        assert "Index#" in headers
        assert "Date" in headers
        assert "Name" in headers
        assert "Status" in headers, "Status from file2 should be preserved"
        assert "Category" in headers, "Category from file3 should be preserved"
        assert "Priority" in headers, "Priority from file3 should be preserved"

    def test_column_order_template_first_then_new(self):
        """Test that template columns appear first, followed by new columns."""
        # Create template with specific column order
        template_fixture = create_excel_with_basic_columns()  # [Index#, Date, Name]
        template_path = template_fixture["path"]

        # Create DataFrame with new columns that would appear differently if sorted
        df = pd.DataFrame(
            {
                "Index#": ["1", "2"],
                "Date": ["2024-01-01", "2024-01-02"],
                "Name": ["Doc A", "Doc B"],
                "Zebra": ["Z1", "Z2"],  # Would be last if alphabetically sorted
                "Apple": ["A1", "A2"],  # Would be first if alphabetically sorted
            }
        )

        # Save with add_missing_columns=True
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as output_file:
            output_path = Path(output_file.name)

            self.repo.save(
                df=df,
                template_path=str(template_path),
                target_sheet="Index",
                out_path=str(output_path),
                add_missing_columns=True,
            )

        # Load result and verify column order
        wb = load_workbook(str(output_path))
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

        # Clean up
        template_path.unlink()
        output_path.unlink()
        wb.close()

        # Assert template columns come first, then new columns in order they were added
        assert headers[0] == "Index#", "Template columns should come first"
        assert headers[1] == "Date", "Template columns should maintain order"
        assert headers[2] == "Name", "Template columns should maintain order"
        # New columns should be after template columns (not alphabetically sorted)
        assert headers[3] in ["Zebra", "Apple"], "New columns should be appended"
        assert headers[4] in ["Zebra", "Apple"], "New columns should be appended"
        assert len(headers) == 5
