#!/usr/bin/env python3
"""
Smoke/Parity tests for ExcelOpenpyxlRepo save behavior using an existing fixture.

This test verifies that after applying pure transforms and saving back into a
template workbook/sheet, the resulting workbook data matches the transformed
DataFrame (content-identical on intersecting columns).
"""

import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from adapters.excel_repo import ExcelOpenpyxlRepo
from core.transform.excel import sort_and_renumber


def _first_sheet_name(xlsx_path: str) -> str:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        return wb.sheetnames[0]
    finally:
        wb.close()


def test_excel_save_parity_on_fixture():
    # Create a minimal test Excel file in memory
    test_data = {
        "Index#": ["1", "2", "3"],
        "Document Type": ["Assignment", "Report", "Contract"],
        "Document Date": ["2024-01-01", "2024-01-02", "2024-01-03"],
        "Received Date": ["2024-01-15", "2024-01-16", "2024-01-17"],
    }
    df_original = pd.DataFrame(test_data)

    # Create temporary fixture file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_fixture:
        fixture_path = Path(tmp_fixture.name)
        df_original.to_excel(fixture_path, index=False, sheet_name="Index")

    sheet = "Index"

    # Load fixture as strings to mimic app behavior
    df = pd.read_excel(fixture_path, dtype=str, sheet_name=sheet)

    # Apply pure transforms (minimal data cleaning for test)
    # Clean Index# column like clean_types() used to do
    if "Index#" in df.columns:
        df["Index#"] = df["Index#"].astype(str).str.strip().replace("nan", "")

    df_expected = sort_and_renumber(df)

    # Save into a temp output using the same template and target sheet
    repo = ExcelOpenpyxlRepo()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = Path(tmp.name)

    # Use the original file as template to preserve formatting/sheets
    repo.save(
        df_expected,
        template_path=str(fixture_path),
        target_sheet=sheet,
        out_path=str(out_path),
    )

    # Read workbook back and compare on intersecting columns
    df_out = pd.read_excel(out_path, dtype=str, sheet_name=sheet)

    common_cols = [
        c
        for c in df_expected.columns
        if c in df_out.columns and c != "Bookmark Formula"
    ]
    assert common_cols, (
        "No common columns between expected DataFrame and saved workbook"
    )

    # Align and compare content (string-wise), normalize dates (strip time part)
    left = df_expected[common_cols].astype(str).reset_index(drop=True)
    right = df_out[common_cols].astype(str).reset_index(drop=True)
    # Normalize NaN-like strings to empty and strip time from dates
    left = left.map(lambda s: "" if str(s).lower() == "nan" else s)
    right = right.map(lambda s: "" if str(s).lower() == "nan" else s)
    for date_col in ("Document Date", "Received Date"):
        if date_col in left.columns and date_col in right.columns:
            left[date_col] = left[date_col].map(lambda s: s.split(" ")[0])
            right[date_col] = right[date_col].map(lambda s: s.split(" ")[0])
    pd.testing.assert_frame_equal(left, right, check_dtype=False)

    # Cleanup temporary files
    fixture_path.unlink(missing_ok=True)
    out_path.unlink(missing_ok=True)
