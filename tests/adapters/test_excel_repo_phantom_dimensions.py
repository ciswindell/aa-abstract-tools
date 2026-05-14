"""Regression tests for the save loops in adapters/excel_repo.py.

Two invariants:

1. **Phantom dimensions must not cause a hang.** Some real-world templates
   report `ws.max_row` near Excel's hard limit (~1,048,576) even when only a
   handful of rows actually contain data. The save loops must not iterate
   over every phantom cell.

2. **Template rows beyond the (smaller) DataFrame must still be cleared.**
   The filter pipeline produces a DataFrame smaller than the template's row
   count. The save must blank out the now-orphaned rows so they don't leak
   into the output. This is the load-bearing reason the loops exist.

A correct fix has to satisfy both — bounding strictly to `len(df)+1` would
satisfy (1) but break (2).
"""

import signal
import time

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from adapters.excel_repo import ExcelOpenpyxlRepo


class _TimeoutError(Exception):
    pass


def _on_alarm(signum, frame):
    raise _TimeoutError("save() exceeded budget — phantom dimensions not bounded")


def _build_template_with_phantom_dimensions(
    path, sheet_name, phantom_row=500_000, phantom_col=16
):
    """Create an .xlsx whose used-range extends to (phantom_row, phantom_col)
    while real data is only 4 rows x 4 cols."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ["Index#", "Document Type", "Document Date", "Received Date"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r in range(2, 5):
        ws.cell(row=r, column=1, value=str(r - 1))
        ws.cell(row=r, column=2, value="Release")
        ws.cell(row=r, column=3, value="2024-01-01")
        ws.cell(row=r, column=4, value="2024-01-02")
    # Phantom: a real value far outside the data region expands max_row/max_column.
    ws.cell(row=phantom_row, column=phantom_col, value=" ")
    wb.save(path)
    wb.close()


def test_save_completes_quickly_on_workbook_with_phantom_dimensions(tmp_path):
    sheet_name = "Index"
    template = tmp_path / "phantom_template.xlsx"
    out_path = tmp_path / "out.xlsx"

    _build_template_with_phantom_dimensions(template, sheet_name)

    # Sanity-check the fixture actually reproduces phantom dimensions.
    check = load_workbook(template)
    assert check[sheet_name].max_row >= 500_000, "fixture didn't extend dimensions"
    assert check[sheet_name].max_column >= 16, "fixture didn't extend columns"
    check.close()

    df = pd.DataFrame(
        {
            "Index#": ["1", "2", "3"],
            "Document Type": ["Release", "Release", "Release"],
            "Document Date": ["2024-01-01", "2024-01-01", "2024-01-01"],
            "Received Date": ["2024-01-02", "2024-01-02", "2024-01-02"],
        }
    )

    repo = ExcelOpenpyxlRepo()

    # Budget: 5 s. With the bug a 500 000-row phantom range causes both the
    # value-clearing and fill-clearing loops to run for many seconds each.
    # Hard timeout via SIGALRM prevents the failing test itself from hanging CI.
    signal.signal(signal.SIGALRM, _on_alarm)
    signal.alarm(10)
    try:
        t0 = time.monotonic()
        try:
            repo.save(df, str(template), sheet_name, str(out_path))
        except _TimeoutError:
            pytest.fail(
                "save() hung past the 10 s hard limit on a phantom-dimension workbook"
            )
        elapsed = time.monotonic() - t0
    finally:
        signal.alarm(0)

    assert elapsed < 5.0, (
        f"save() took {elapsed:.1f}s on a workbook with phantom dimensions — "
        "fill-clearing loop is iterating over the phantom range"
    )


def _build_template_with_n_data_rows(path, sheet_name, n):
    """Create a template with exactly `n` populated data rows after the header."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ["Index#", "Document Type", "Document Date", "Received Date"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r in range(2, n + 2):
        ws.cell(row=r, column=1, value=str(r - 1))
        ws.cell(row=r, column=2, value="OldType")
        ws.cell(row=r, column=3, value="2020-01-01")
        ws.cell(row=r, column=4, value="2020-01-02")
    wb.save(path)
    wb.close()


def test_save_blanks_template_rows_beyond_smaller_dataframe(tmp_path):
    """When the DataFrame has fewer rows than the template (filter pipeline),
    cells in template rows past `len(df)` must be blanked in the output.
    """
    sheet_name = "Index"
    template = tmp_path / "10row_template.xlsx"
    out_path = tmp_path / "out.xlsx"

    # Template: 10 rows of real data. DataFrame: 4 rows (simulating a filter).
    _build_template_with_n_data_rows(template, sheet_name, n=10)

    df = pd.DataFrame(
        {
            "Index#": ["1", "2", "3", "4"],
            "Document Type": ["NewType", "NewType", "NewType", "NewType"],
            "Document Date": ["2024-01-01"] * 4,
            "Received Date": ["2024-01-02"] * 4,
        }
    )

    ExcelOpenpyxlRepo().save(df, str(template), sheet_name, str(out_path))

    out = load_workbook(out_path)[sheet_name]
    # Rows 2..5 have the new values
    for r in range(2, 6):
        assert out.cell(r, 2).value == "NewType", f"row {r} col 2 not updated"
    # Rows 6..11 must be blank — they were real template rows that fell
    # outside the smaller DataFrame, and must not leak into the output.
    for r in range(6, 12):
        v = out.cell(r, 2).value
        assert v in (None, ""), (
            f"row {r} col 2 leaked stale template value {v!r} — "
            "value-clear loop didn't extend past len(df)"
        )
