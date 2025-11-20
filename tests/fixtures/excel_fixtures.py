#!/usr/bin/env python3
"""
Test fixtures for Excel files with varying column structures.
Used to test column preservation during merge operations.
"""

import tempfile
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook


def create_test_excel(columns: List[str], data: List[List[Any]] = None, sheet_name: str = "Index") -> Path:
    """
    Create a test Excel file with specified columns and data.
    
    Args:
        columns: List of column header names
        data: Optional list of rows (each row is a list of values)
        sheet_name: Name of the worksheet
        
    Returns:
        Path to the created temporary Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Add headers
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Add data if provided
    if data:
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(temp_file.name)
    wb.close()
    
    return Path(temp_file.name)


def create_excel_with_basic_columns() -> Dict[str, Any]:
    """
    Create a basic Excel file with columns [A, B, C].
    
    Returns:
        Dict with 'path' and 'columns' keys
    """
    columns = ["Index#", "Date", "Name"]
    data = [
        ["1", "2024-01-01", "Document A"],
        ["2", "2024-01-02", "Document B"],
        ["3", "2024-01-03", "Document C"],
    ]
    
    path = create_test_excel(columns, data)
    return {"path": path, "columns": columns, "row_count": len(data)}


def create_excel_with_extra_columns() -> Dict[str, Any]:
    """
    Create an Excel file with columns [A, B, C, D, E] for merge testing.
    
    Returns:
        Dict with 'path' and 'columns' keys
    """
    columns = ["Index#", "Date", "Name", "Status", "Comments"]
    data = [
        ["1", "2024-02-01", "Document D", "Approved", "Looks good"],
        ["2", "2024-02-02", "Document E", "Pending", "Needs review"],
    ]
    
    path = create_test_excel(columns, data)
    return {"path": path, "columns": columns, "row_count": len(data)}


def create_excel_with_disjoint_columns() -> Dict[str, Any]:
    """
    Create an Excel file with completely different columns for testing disjoint sets.
    
    Returns:
        Dict with 'path' and 'columns' keys
    """
    columns = ["Index#", "Priority", "Category", "Owner"]
    data = [
        ["1", "High", "Security", "Alice"],
        ["2", "Low", "Performance", "Bob"],
    ]
    
    path = create_test_excel(columns, data)
    return {"path": path, "columns": columns, "row_count": len(data)}


def create_excel_with_case_variations() -> Dict[str, Any]:
    """
    Create an Excel file with column names that vary in case.
    
    Returns:
        Dict with 'path' and 'columns' keys
    """
    columns = ["INDEX#", "date", "Name"]
    data = [
        ["1", "2024-03-01", "Document F"],
    ]
    
    path = create_test_excel(columns, data)
    return {"path": path, "columns": columns, "row_count": len(data)}


def create_excel_with_whitespace_variations() -> Dict[str, Any]:
    """
    Create an Excel file with column names that have varying whitespace.
    
    Returns:
        Dict with 'path' and 'columns' keys
    """
    columns = ["Index#", " Date ", "Name  "]
    data = [
        ["1", "2024-04-01", "Document G"],
    ]
    
    path = create_test_excel(columns, data)
    return {"path": path, "columns": columns, "row_count": len(data)}


def create_excel_with_system_columns() -> Dict[str, Any]:
    """
    Create an Excel file with both user and system columns.
    System columns should be excluded from merge output.
    
    Returns:
        Dict with 'path' and 'columns' keys
    """
    columns = ["Index#", "Date", "Name", "_include", "Document_ID", "_original_index"]
    data = [
        ["1", "2024-05-01", "Document H", True, "file1_001", 0],
        ["2", "2024-05-02", "Document I", False, "file1_002", 1],
    ]
    
    path = create_test_excel(columns, data)
    return {"path": path, "columns": columns, "row_count": len(data)}

