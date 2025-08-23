#!/usr/bin/env python3
"""
Abstract Renumber Tool
A class-based application for sorting Excel data and renumbering PDF bookmarks.
"""

import os
import shutil
import tkinter as tk
from datetime import datetime
from pathlib import Path
import tempfile
from tkinter import filedialog, messagebox, ttk
from typing import List, Optional, Tuple


from openpyxl import load_workbook

# Import our modular classes
from excel_processor import ExcelProcessor
from pdf_processor import PDFProcessor
from pdf_validator import PDFValidator
from validators.input_sheet_validator import validate as validate_sheet

# New architecture imports
from adapters.excel_repo import ExcelOpenpyxlRepo
from adapters.pdf_repo import PdfPyPDF2Repo
from adapters.logger_tk import TkLogger
from core.models import Options
from core.services.validate import ValidationService
from core.services.renumber import RenumberService


class AbstractRenumberGUI:
    """Handles the GUI components and user interactions."""

    def __init__(self, root: tk.Tk, controller: "AbstractRenumberTool") -> None:
        self.root = root
        self.controller = controller
        self.excel_file: Optional[str] = None
        self.pdf_file: Optional[str] = None
        self.backup_enabled: tk.BooleanVar = tk.BooleanVar(
            value=True
        )  # Backup enabled by default
        self.sort_bookmarks_enabled: tk.BooleanVar = tk.BooleanVar(
            value=False
        )  # Sort bookmarks disabled by default
        self.reorder_pages_enabled: tk.BooleanVar = tk.BooleanVar(
            value=False
        )  # Reorder pages disabled by default

        # Initialize GUI components
        self.excel_label: Optional[ttk.Label] = None
        self.pdf_label: Optional[ttk.Label] = None
        self.process_button: Optional[ttk.Button] = None
        self.status_text: Optional[tk.Text] = None
        self.backup_checkbox: Optional[ttk.Checkbutton] = None
        self.backup_info_label: Optional[ttk.Label] = None
        self.sort_bookmarks_checkbox: Optional[ttk.Checkbutton] = None
        self.reorder_pages_checkbox: Optional[ttk.Checkbutton] = None
        self.reorder_pages_info_label: Optional[ttk.Label] = None

        self.setup_window()
        self.setup_gui()

    def setup_window(self) -> None:
        """Setup main window properties."""
        self.root.title("Abstract Renumber Tool")
        self.root.geometry("600x500")
        self.root.resizable(True, True)

        # Center window
        self._center_window()

        # Configure grid weights
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

    def _center_window(self) -> None:
        """Center the main window on the screen."""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        pos_x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        pos_y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{pos_x}+{pos_y}")

    def setup_gui(self) -> None:
        """Initialize GUI components."""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        main_frame.grid_rowconfigure(6, weight=1)  # Updated for new row
        main_frame.grid_columnconfigure(1, weight=1)

        # Title
        title_label = ttk.Label(
            main_frame, text="Abstract Renumber Tool", font=("Arial", 16, "bold")
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))

        # File selection
        self._create_file_selection(main_frame)

        # Backup options
        self._create_backup_options(main_frame)

        # Process button
        self.process_button = ttk.Button(
            main_frame,
            text="Process Files",
            command=self.controller.process_files,
            state="disabled",
        )
        self.process_button.grid(row=5, column=0, columnspan=2, pady=20)

        # Status area
        self._create_status_area(main_frame)

        # Initial status
        self.log_status("Ready. Please select Excel and PDF files.")

    def _create_file_selection(self, parent: ttk.Frame) -> None:
        """Create file selection components."""
        # Section header
        ttk.Label(parent, text="Select Files:", font=("Arial", 12, "bold")).grid(
            row=1, column=0, columnspan=2, sticky=tk.W, pady=(0, 10)
        )

        # Excel file
        ttk.Label(parent, text="Excel File:").grid(row=2, column=0, sticky=tk.W, pady=5)
        excel_frame = ttk.Frame(parent)
        excel_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        excel_frame.grid_columnconfigure(1, weight=1)

        ttk.Button(excel_frame, text="Browse...", command=self._select_excel_file).grid(
            row=0, column=0, sticky=tk.W
        )
        self.excel_label = ttk.Label(
            excel_frame, text="No file selected", foreground="gray"
        )
        self.excel_label.grid(row=0, column=1, sticky=tk.W, padx=(10, 0))

        # PDF file
        ttk.Label(parent, text="PDF File:").grid(row=3, column=0, sticky=tk.W, pady=5)
        pdf_frame = ttk.Frame(parent)
        pdf_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5, padx=(10, 0))
        pdf_frame.grid_columnconfigure(1, weight=1)

        ttk.Button(pdf_frame, text="Browse...", command=self._select_pdf_file).grid(
            row=0, column=0, sticky=tk.W
        )
        self.pdf_label = ttk.Label(
            pdf_frame, text="No file selected", foreground="gray"
        )
        self.pdf_label.grid(row=0, column=1, sticky=tk.W, padx=(10, 0))

    def _create_backup_options(self, parent: ttk.Frame) -> None:
        """Create backup options section with checkbox."""
        # Options frame
        options_frame = ttk.LabelFrame(parent, text="Processing Options", padding="10")
        options_frame.grid(
            row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0)
        )
        options_frame.grid_columnconfigure(0, weight=1)

        # Backup checkbox with description
        backup_frame = ttk.Frame(options_frame)
        backup_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5)
        backup_frame.grid_columnconfigure(1, weight=1)

        self.backup_checkbox = ttk.Checkbutton(
            backup_frame,
            text="Create backup files before processing",
            variable=self.backup_enabled,
            command=self._on_backup_option_changed,
        )
        self.backup_checkbox.grid(row=0, column=0, sticky=tk.W)

        # Info label for backup option
        self.backup_info_label = ttk.Label(
            backup_frame,
            text="✅ Recommended: Creates timestamped backups for safety",
            foreground="green",
            font=("Arial", 9),
        )
        self.backup_info_label.grid(
            row=1, column=0, sticky=tk.W, padx=(20, 0), pady=(2, 0)
        )

        # Sort PDF Bookmarks checkbox
        sort_bookmarks_frame = ttk.Frame(options_frame)
        sort_bookmarks_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(15, 5))
        sort_bookmarks_frame.grid_columnconfigure(1, weight=1)

        self.sort_bookmarks_checkbox = ttk.Checkbutton(
            sort_bookmarks_frame,
            text="Sort PDF Bookmarks",
            variable=self.sort_bookmarks_enabled,
            command=self._on_sort_bookmarks_option_changed,
        )
        self.sort_bookmarks_checkbox.grid(row=0, column=0, sticky=tk.W)

        # Reorder Pages checkbox (initially disabled)
        reorder_pages_frame = ttk.Frame(options_frame)
        reorder_pages_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5)
        reorder_pages_frame.grid_columnconfigure(1, weight=1)

        self.reorder_pages_checkbox = ttk.Checkbutton(
            reorder_pages_frame,
            text="Reorder Pages to Match Bookmarks",
            variable=self.reorder_pages_enabled,
            state="disabled",  # Initially disabled
        )
        self.reorder_pages_checkbox.grid(row=0, column=0, sticky=tk.W, padx=(20, 0))

        # Info label for reorder pages option
        self.reorder_pages_info_label = ttk.Label(
            reorder_pages_frame,
            text="Requires 'Sort PDF Bookmarks' to be enabled",
            foreground="gray",
            font=("Arial", 9),
        )
        self.reorder_pages_info_label.grid(
            row=1, column=0, sticky=tk.W, padx=(40, 0), pady=(2, 0)
        )

    def _on_backup_option_changed(self) -> None:
        """Handle changes to the backup option checkbox."""
        if self.backup_enabled.get():
            self.backup_info_label.config(
                text="✅ Recommended: Creates timestamped backups for safety",
                foreground="green",
            )
        else:
            self.backup_info_label.config(
                text="⚠️  Warning: Original files will be overwritten directly",
                foreground="orange",
            )

    def _on_sort_bookmarks_option_changed(self) -> None:
        """Handle changes to the sort bookmarks option checkbox."""
        if self.sort_bookmarks_enabled.get():
            # Enable and check the reorder pages checkbox when sort bookmarks is checked
            self.reorder_pages_checkbox.config(state="normal")
            self.reorder_pages_enabled.set(True)  # Auto-check by default
            self.reorder_pages_info_label.config(
                text="Optional: Physically reorder pages to match sorted bookmarks",
                foreground="green",
            )
        else:
            # Disable and uncheck the reorder pages checkbox when sort bookmarks is unchecked
            self.reorder_pages_enabled.set(False)
            self.reorder_pages_checkbox.config(state="disabled")
            self.reorder_pages_info_label.config(
                text="Requires 'Sort PDF Bookmarks' to be enabled", foreground="gray"
            )

    def _create_status_area(self, parent: ttk.Frame) -> None:
        """Create status text area with scrollbar."""
        status_frame = ttk.LabelFrame(parent, text="Status", padding="10")
        status_frame.grid(
            row=6,
            column=0,
            columnspan=2,
            sticky=(tk.W, tk.E, tk.N, tk.S),
            pady=(20, 0),  # Updated row number
        )
        status_frame.grid_rowconfigure(0, weight=1)
        status_frame.grid_columnconfigure(0, weight=1)

        self.status_text = tk.Text(status_frame, height=8, width=70, wrap=tk.WORD)
        self.status_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        scrollbar = ttk.Scrollbar(
            status_frame, orient="vertical", command=self.status_text.yview
        )
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.status_text.configure(yscrollcommand=scrollbar.set)

    def _get_default_directory(self) -> str:
        """Get the user's Downloads directory."""
        downloads_path = Path.home() / "Downloads"
        return str(downloads_path) if downloads_path.exists() else str(Path.home())

    def _select_excel_file(self) -> None:
        """Handle Excel file selection."""
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            initialdir=self._get_default_directory(),
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )

        if file_path:
            self.excel_file = file_path
            self.excel_label.config(
                text=os.path.basename(file_path), foreground="black"
            )
            self._check_files_ready()

    def _select_pdf_file(self) -> None:
        """Handle PDF file selection."""
        file_path = filedialog.askopenfilename(
            title="Select PDF File",
            initialdir=self._get_default_directory(),
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )

        if file_path:
            self.pdf_file = file_path
            self.pdf_label.config(text=os.path.basename(file_path), foreground="black")
            self._check_files_ready()

    def _check_files_ready(self) -> None:
        """Enable process button if both files are selected."""
        if self.excel_file and self.pdf_file:
            self.process_button.config(state="normal")
            self.log_status("Ready to process!")
        else:
            self.process_button.config(state="disabled")

    def log_status(self, message: str) -> None:
        """Add a timestamped message to the status text."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.root.update()

    def get_selected_files(self) -> Tuple[Optional[str], Optional[str]]:
        """Return the selected file paths."""
        return self.excel_file, self.pdf_file

    def get_backup_enabled(self) -> bool:
        """Return whether backup creation is enabled."""
        return self.backup_enabled.get()

    def get_sort_bookmarks_enabled(self) -> bool:
        """Return whether PDF bookmark sorting is enabled."""
        return self.sort_bookmarks_enabled.get()

    def get_reorder_pages_enabled(self) -> bool:
        """Return whether page reordering is enabled."""
        return self.reorder_pages_enabled.get()


class AbstractRenumberTool:
    """Main application controller that orchestrates the components."""

    def __init__(self):
        """Initialize the Abstract Renumber Tool with required configuration."""
        # Configuration
        self.required_columns = [
            "Index#",
            "Document Type",
            "Legal Description",
            "Grantee",
            "Grantor",
            "Document Date",
            "Received Date",
        ]

        # Configured processing sheet name (case-insensitive match)
        self.processing_sheet_name: Optional[str] = "Index"
        self.selected_processing_sheet_name: Optional[str] = None

        # Initialize components
        self.root = tk.Tk()
        self.excel_processor = ExcelProcessor(self.required_columns)
        self.gui = AbstractRenumberGUI(self.root, self)
        self.pdf_processor: Optional[PDFProcessor] = None

    def get_processing_sheet_name(self) -> Optional[str]:
        """Return the configured processing sheet name if any."""
        return self.processing_sheet_name

    def process_files(self):
        """Main processing function that orchestrates processing via RenumberService."""
        try:
            excel_file, pdf_file = self.gui.get_selected_files()

            # Resolve/select processing sheet name to match legacy UX
            target_sheet = self._resolve_processing_sheet_name(excel_file)
            if target_sheet is None:
                target_sheet = self._prompt_user_select_sheet(excel_file)
                if target_sheet is None:
                    return
            self.selected_processing_sheet_name = target_sheet

            # Backups (preserve legacy behavior)
            backup_enabled = self.gui.get_backup_enabled()
            if backup_enabled:
                self._create_backup_files(excel_file, pdf_file)

            # Build service and options
            logger = TkLogger(self.gui.log_status)
            excel_repo = ExcelOpenpyxlRepo()
            pdf_repo = PdfPyPDF2Repo()
            validator = ValidationService(self.required_columns)
            service = RenumberService(excel_repo, pdf_repo, validator, logger)

            opts = Options(
                backup=backup_enabled,
                sort_bookmarks=self.gui.get_sort_bookmarks_enabled(),
                reorder_pages=self.gui.get_reorder_pages_enabled(),
                sheet_name=target_sheet,
            )

            # Run service
            result = service.run(excel_file, pdf_file, opts)
            if not result.success:
                raise RuntimeError(result.message or "Unknown error")

            # Completion (preserve simple success dialog)
            self._show_completion_success(excel_file, pdf_file)

        except (ValueError, OSError, RuntimeError) as e:
            self.gui.log_status(f"Error: {str(e)}")
            messagebox.showerror("Processing Error", f"Processing failed: {str(e)}")

    def _process_excel_file(self, file_path: str) -> bool:
        """Process Excel file with validation and mapping."""
        try:
            # Load file
            target_sheet = self._resolve_processing_sheet_name(file_path)
            if target_sheet is None:
                # Prompt user to select a sheet
                target_sheet = self._prompt_user_select_sheet(file_path)
                if target_sheet is None:
                    return False
            # Remember the chosen sheet for save phase
            self.selected_processing_sheet_name = target_sheet
            self.excel_processor.load_file(file_path, sheet_name=target_sheet)

            # Get file info
            info = self.excel_processor.get_column_info()

            # Check for duplicate columns
            duplicates = self.excel_processor.check_duplicate_columns()
            if duplicates:
                messagebox.showwarning(
                    "Duplicate Columns",
                    "The Excel file has duplicate column names:\n"
                    + "\n".join(f"• {col}" for col in duplicates)
                    + "\n\nThis may cause issues.",
                )

            # Validate columns via validator (case-insensitive, required-only duplicates)
            results = validate_sheet(
                self.excel_processor.get_dataframe(), self.required_columns
            )
            missing_columns = results["missing"]
            required_dupes = results["duplicates"]
            if missing_columns or required_dupes:
                details = []
                if missing_columns:
                    details.append(
                        "Missing required columns:\n"
                        + "\n".join(f"• {c}" for c in missing_columns)
                    )
                if required_dupes:
                    details.append(
                        "Duplicate required columns:\n"
                        + "\n".join(f"• {c}" for c in required_dupes)
                    )
                messagebox.showerror("Column Validation Error", "\n\n".join(details))
                return False

            return True

        except FileNotFoundError as e:
            self._handle_error("File Not Found", str(e))
            return False

    def _prompt_user_select_sheet(self, file_path: str) -> Optional[str]:
        """Show a dropdown (readonly) of sheet names to select from."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()

            if not names:
                return None

            # Determine default selection (prefer configured 'Index')
            desired = (self.get_processing_sheet_name() or "").lower()
            default_index = 0
            for i, n in enumerate(names):
                if n.lower() == desired:
                    default_index = i
                    break

            # Modal dialog with dropdown
            self.gui.log_status("Select processing sheet...")
            self.root.update()

            dialog = tk.Toplevel(self.root)
            dialog.title("Select Sheet")
            dialog.transient(self.root)
            dialog.grab_set()

            expected = self.get_processing_sheet_name() or "Index"
            msg = (
                f"Warning: The Excel file does not contain a worksheet named '{expected}'.\n\n"
                "Please select the worksheet that needs to be processed."
            )
            ttk.Label(dialog, text=msg, wraplength=420, justify=tk.LEFT).grid(
                row=0, column=0, columnspan=2, padx=12, pady=(12, 6), sticky=tk.W
            )

            selected_var = tk.StringVar(value=names[default_index])
            combo = ttk.Combobox(
                dialog,
                state="readonly",
                values=names,
                textvariable=selected_var,
                width=40,
            )
            combo.grid(
                row=1,
                column=0,
                columnspan=2,
                padx=12,
                pady=(0, 12),
                sticky=(tk.W, tk.E),
            )
            combo.current(default_index)

            chosen = {"value": None}

            def on_ok():
                chosen["value"] = selected_var.get()
                dialog.destroy()

            def on_cancel():
                chosen["value"] = None
                dialog.destroy()

            ok_btn = ttk.Button(dialog, text="OK", command=on_ok)
            cancel_btn = ttk.Button(dialog, text="Cancel", command=on_cancel)
            ok_btn.grid(row=2, column=0, padx=(12, 6), pady=(0, 12), sticky=tk.E)
            cancel_btn.grid(row=2, column=1, padx=(6, 12), pady=(0, 12), sticky=tk.W)

            dialog.bind("<Return>", lambda _: on_ok())
            dialog.bind("<Escape>", lambda _: on_cancel())
            dialog.wait_window(dialog)

            return chosen["value"]
        except PermissionError as e:
            self._handle_error(
                "Permission Error",
                f"Cannot access Excel file:\n{str(e)}\n\n"
                "Please check file permissions and ensure the file is not open.",
            )
            return False
        except ValueError as e:
            self._handle_error("Invalid File", str(e))
            return False
        except OSError as e:
            self._handle_error(
                "Error",
                f"Failed to load Excel file:\n{str(e)}\n\n"
                "Please ensure the file is a valid Excel file.",
            )
            return False
        except Exception as e:
            self._handle_error("Sheet Selection Error", str(e))
            return None

    def _resolve_processing_sheet_name(self, file_path: str) -> Optional[str]:
        """Resolve the processing sheet name, case-insensitively defaulting to 'Index'.

        If the configured sheet name is not found, return None to trigger GUI picker later.
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            desired = (self.get_processing_sheet_name() or "").lower()
            names = wb.sheetnames
            lower_to_name = {n.lower(): n for n in names}
            wb.close()
            if desired and desired in lower_to_name:
                return lower_to_name[desired]
            return None
        except Exception:
            return None

    def _handle_missing_columns(
        self, missing_columns: List[str], available_columns: List[str]
    ) -> bool:
        """Handle missing columns by showing an error and aborting (no mapping)."""
        message = "Missing required columns:\n" + "\n".join(
            f"• {c}" for c in missing_columns
        )
        self._handle_error("Missing Columns", message)
        return False

    def _validate_pdf_bookmarks(self, pdf_file: str) -> bool:
        """Validate PDF bookmarks for conflicts before processing.

        Args:
            pdf_file: Path to the PDF file to validate

        Returns:
            bool: True if validation passes, False if conflicts found
        """
        try:
            self.gui.log_status("Validating PDF bookmarks...")

            # Use existing PDFProcessor to extract bookmarks (DRY principle)
            temp_pdf_processor = PDFProcessor()
            temp_pdf_processor.load_pdf(pdf_file)
            bookmarks = temp_pdf_processor.bookmarks

            # Validate bookmarks using PDFValidator (Single Responsibility)
            conflicts = PDFValidator.validate_bookmark_page_conflicts(bookmarks)

            if conflicts:
                # Format error message with conflict details
                conflict_details = []
                for conflict in conflicts["conflicts"]:
                    page = conflict["page"]
                    bookmark_titles = conflict["bookmarks"]
                    conflict_details.append(
                        f"Page {page}: {', '.join(bookmark_titles)}"
                    )

                error_message = (
                    "PDF bookmark conflicts detected!\n\n"
                    "Multiple bookmarks point to the same page(s):\n"
                    + "\n".join(conflict_details)
                    + "\n\nPlease fix these conflicts before processing."
                )

                self._handle_error("Bookmark Conflicts", error_message)
                return False

            self.gui.log_status("PDF bookmark validation passed.")
            return True

        except (ValueError, OSError, RuntimeError) as e:
            self._handle_error(
                "Validation Error", f"Failed to validate PDF bookmarks: {str(e)}"
            )
            return False

    def _handle_error(self, title: str, message: str):
        """Handle errors with consistent logging and user feedback."""
        self.gui.log_status(f"Error: {message}")
        messagebox.showerror(title, message)

    def _perform_excel_processing(self) -> bool:
        """Perform Excel file processing after validation."""
        try:
            # Get source file paths
            excel_file, pdf_file = self.gui.get_selected_files()
            backup_enabled = self.gui.get_backup_enabled()

            # Basic file validation
            if not os.path.exists(excel_file) or not os.path.exists(pdf_file):
                messagebox.showerror("File Error", "Source files not found")
                return False

            # Create backups if enabled
            if backup_enabled:
                self._create_backup_files(excel_file, pdf_file)

            # Generate output filenames
            excel_output_path, pdf_output_path = self._generate_output_filenames(
                excel_file, pdf_file
            )

            # Save Excel using atomic temp-file then rename to final
            self._save_excel_atomically(excel_file, excel_output_path)

            # Save PDF
            self._save_updated_pdf_file(pdf_output_path)

            # Success
            self._show_completion_success(excel_output_path, pdf_output_path)
            return True

        except (ValueError, OSError, RuntimeError) as e:
            self.gui.log_status(f"Error: {str(e)}")
            messagebox.showerror("Processing Error", f"Processing failed: {str(e)}")
            return False

    def _show_completion_success(
        self, excel_output_path: str, pdf_output_path: str  # noqa: ARG002
    ) -> None:
        """Show simple completion success message."""
        self.gui.log_status("Complete!")

        # Simple success dialog
        message = (
            f"Files processed successfully!\n\n"
            f"Saved to: {os.path.dirname(excel_output_path)}"
        )
        messagebox.showinfo("Processing Complete", message)

    def _save_processed_excel_file(self, excel_output_path: str) -> bool:
        """Save the processed Excel file with basic error handling."""
        try:
            # Use the resolved/selected sheet name; fail if not set
            excel_file, _ = self.gui.get_selected_files()
            final_sheet = (
                self.selected_processing_sheet_name
                or self._resolve_processing_sheet_name(excel_file)
            )
            if not final_sheet:
                raise RuntimeError("Processing sheet not selected or found")
            # Adapter-driven save now handles formulas via template
            ExcelOpenpyxlRepo().save(
                self.excel_processor.get_dataframe(),
                template_path=excel_file,
                target_sheet=final_sheet,
                out_path=excel_output_path,
            )
            return True
        except (ValueError, OSError, RuntimeError) as e:
            raise RuntimeError(f"Failed to save Excel file: {str(e)}") from e

    def _save_excel_atomically(
        self, input_excel_path: str, final_output_path: str
    ) -> None:
        """Write Excel output atomically using a temp file then rename.

        1. Copy input workbook to a temp file in the destination directory.
        2. Write changes into the temp file (preserving formatting).
        3. Atomically replace/move temp file to final output path.
        4. Clean up temp file on any error.
        """
        dest_dir = os.path.dirname(final_output_path) or "."
        temp_file = None
        try:
            # Create temp file path in same directory to ensure atomic rename
            with tempfile.NamedTemporaryFile(
                prefix=".tmp_excel_", suffix=".xlsx", dir=dest_dir, delete=False
            ) as tf:
                temp_file = tf.name

            # Copy input workbook to temp (template preservation)
            shutil.copy2(input_excel_path, temp_file)

            # Determine processing sheet
            excel_file, _ = self.gui.get_selected_files()
            final_sheet = (
                self.selected_processing_sheet_name
                or self._resolve_processing_sheet_name(excel_file)
            )
            if not final_sheet:
                raise RuntimeError("Processing sheet not selected or found")

            # Write values and formatting into temp workbook
            ExcelOpenpyxlRepo().save(
                self.excel_processor.get_dataframe(),
                template_path=input_excel_path,
                target_sheet=final_sheet,
                out_path=temp_file,
            )

            # Atomically replace/move temp file to final output
            os.replace(temp_file, final_output_path)
            temp_file = None  # Prevent cleanup
        finally:
            # Ensure no temp remnants
            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except Exception:
                    pass

    def _save_updated_pdf_file(self, pdf_output_path: str) -> bool:
        """Save the updated PDF file with basic error handling."""
        try:
            # Initialize PDFProcessor if needed
            if self.pdf_processor is None:
                self.pdf_processor = PDFProcessor()

            # Load PDF file
            _, pdf_file = self.gui.get_selected_files()
            self.pdf_processor.load_pdf(pdf_file)

            # Copy pages to writer
            self.pdf_processor.copy_pages_to_writer()

            # Get Excel data and index mapping for bookmark generation
            excel_data = self.excel_processor.get_dataframe().to_dict("records")
            index_mapping = self.excel_processor.get_original_index_mapping()

            # Generate new bookmark titles
            new_titles = self.pdf_processor.generate_new_bookmark_titles(
                excel_data, index_mapping
            )

            # Update bookmarks with new titles
            sort_bookmarks = self.gui.get_sort_bookmarks_enabled()
            self.pdf_processor.update_bookmarks_with_new_titles(
                new_titles, sort_naturally=sort_bookmarks
            )

            # Conditional page reordering based on user options
            reorder_pages = self.gui.get_reorder_pages_enabled()

            if reorder_pages:
                self.gui.log_status("Reordering PDF pages to match bookmark order...")
                self.pdf_processor.reorder_pages_by_bookmarks()
                self.gui.log_status("PDF pages reordered successfully.")

            # Save the updated PDF
            self.pdf_processor.save_pdf(pdf_output_path)

            return True

        except (ValueError, OSError, RuntimeError) as e:
            raise RuntimeError(f"Failed to save PDF file: {str(e)}") from e

    def _generate_backup_filename(self, original_path: str) -> str:
        """Generate simple backup filename with timestamp."""
        original_file = Path(original_path)
        file_stem = original_file.stem
        file_suffix = original_file.suffix
        file_dir = original_file.parent

        # Simple timestamp with seconds for uniqueness
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"{file_stem}_backup_{timestamp}{file_suffix}"

        return str(file_dir / backup_filename)

    def _ensure_unique_backup_filename(self, backup_path: str) -> str:
        """Return the backup path as-is since seconds provide sufficient uniqueness."""
        return backup_path

    def _generate_output_filenames(
        self, excel_path: str, pdf_path: str
    ) -> Tuple[str, str]:
        """Generate output filenames using original names (no '_renumbered' suffix)."""
        # New strategy: output files use original filenames
        # The original files will be backed up before processing
        return excel_path, pdf_path

    def _generate_backup_filenames(
        self, excel_path: str, pdf_path: str
    ) -> Tuple[str, str]:
        """Generate backup filenames with timestamp."""
        excel_backup = self._generate_backup_filename(excel_path)
        pdf_backup = self._generate_backup_filename(pdf_path)

        # Ensure uniqueness (now simplified)
        excel_backup = self._ensure_unique_backup_filename(excel_backup)
        pdf_backup = self._ensure_unique_backup_filename(pdf_backup)

        return excel_backup, pdf_backup

    def _create_backup_files(self, excel_path: str, pdf_path: str) -> Tuple[str, str]:
        """Create simple backup copies with timestamp suffix."""
        # Generate backup filenames
        excel_backup_path, pdf_backup_path = self._generate_backup_filenames(
            excel_path, pdf_path
        )

        # Copy files
        shutil.copy2(excel_path, excel_backup_path)
        shutil.copy2(pdf_path, pdf_backup_path)

        return excel_backup_path, pdf_backup_path

    def _prepare_output_excel_as_template(
        self, input_excel_path: str, output_excel_path: str
    ) -> None:
        """Create the output Excel file as a direct copy of the input file.

        This sets up the output workbook to retain all sheets and formatting.
        If the output path equals the input path, no action is taken.
        """
        try:
            if not input_excel_path or not output_excel_path:
                return
            if os.path.abspath(input_excel_path) == os.path.abspath(output_excel_path):
                # No-op if target equals source (naming/backup policy may handle this later)
                return
            # Copy preserving metadata; subsequent steps will modify only intended cells
            shutil.copy2(input_excel_path, output_excel_path)
        except Exception as e:
            raise RuntimeError(
                f"Failed to prepare output Excel template: {str(e)}"
            ) from e

    def run(self) -> None:
        """Start the GUI application."""
        self.root.mainloop()


if __name__ == "__main__":
    app = AbstractRenumberTool()
    app.run()
