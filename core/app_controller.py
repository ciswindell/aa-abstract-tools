#!/usr/bin/env python3
"""
Application controller that orchestrates processing using UI abstraction.
"""

import os
from typing import Optional

from openpyxl import load_workbook

from adapters.excel_repo import ExcelOpenpyxlRepo
from adapters.logger_tk import TkinterLogger
from adapters.pdf_repo import PdfRepo
from core.config import DEFAULT_REQUIRED_COLUMNS, DEFAULT_SHEET_NAME
from core.interfaces import ExcelRepo, Logger, UIController
from core.interfaces import PdfRepo as PdfRepoInterface
from core.services.renumber import RenumberService


class AppController:
    """Application controller that orchestrates processing via UI abstraction."""

    def __init__(
        self,
        ui: UIController,
        excel_repo: Optional[ExcelRepo] = None,
        pdf_repo: Optional[PdfRepoInterface] = None,
        logger: Optional[Logger] = None,
    ) -> None:
        """Initialize with UI controller and optional dependencies."""
        self.ui = ui
        self.required_columns = list(DEFAULT_REQUIRED_COLUMNS)
        self.processing_sheet_name = DEFAULT_SHEET_NAME
        self._excel_repo = excel_repo
        self._pdf_repo = pdf_repo
        self._logger = logger

    def process_files(self) -> None:
        """Main processing function."""
        try:
            excel_file, pdf_file = self.ui.get_file_paths()
            if not excel_file or not pdf_file:
                self.ui.show_error(
                    "Missing Files", "Please select both Excel and PDF files."
                )
                return

            # Resolve processing sheet name
            target_sheet = self._resolve_processing_sheet_name(excel_file)
            if target_sheet is None:
                target_sheet = self._prompt_user_select_sheet(excel_file)
                if target_sheet is None:
                    return

            # Get options from UI
            options = self.ui.get_options()
            options.sheet_name = target_sheet

            # If merging, disable backups regardless of checkbox (originals untouched)
            if options.merge_pairs:
                options.backup = False

            # Build services (use injected or defaults)
            logger = self._logger or TkinterLogger(self.ui)
            excel_repo = self._excel_repo or ExcelOpenpyxlRepo()
            pdf_repo = self._pdf_repo or PdfRepo()

            # Filter prompt will happen after loading/merging in the pipeline
            # This ensures the user sees all merged data when selecting filter values

            # Backups are handled in the service after validation passes.

            # If merging, resolve sheet names for each pair using the same selection flow
            if options.merge_pairs:
                # Always include the primary selection as the first pair
                pairs_with_sheets = [(excel_file, pdf_file, target_sheet)]
                seen = {(excel_file, pdf_file)}

                # Append additional pairs, resolving sheet names per file
                for x_path, p_path in options.merge_pairs or []:
                    if (x_path, p_path) in seen:
                        continue
                    sheet_name = self._resolve_processing_sheet_name(x_path)
                    if sheet_name is None:
                        sheet_name = self._prompt_user_select_sheet(x_path)
                        if sheet_name is None:
                            continue
                    pairs_with_sheets.append((x_path, p_path, sheet_name))
                    seen.add((x_path, p_path))

                options.merge_pairs_with_sheets = pairs_with_sheets

            # Run processing
            service = RenumberService(excel_repo, pdf_repo, logger, self.ui)
            result = service.run(excel_file, pdf_file, options)
            if not result.success:
                raise RuntimeError(result.message or "Unknown error")

            # Show success
            message = f"Files processed successfully!\n\nSaved to: {os.path.dirname(excel_file)}"
            self.ui.show_success(message)

            # Reset GUI for next processing - TEMPORARILY DISABLED FOR DEBUGGING
            # self.ui.reset_gui()

        except (ValueError, OSError, RuntimeError) as e:
            self.ui.show_error("Processing Error", str(e))

    def _resolve_processing_sheet_name(self, file_path: str) -> Optional[str]:
        """Resolve the processing sheet name, case-insensitively."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            desired = (self.processing_sheet_name or "").lower()
            names = wb.sheetnames
            lower_to_name = {n.lower(): n for n in names}
            wb.close()
            if desired and desired in lower_to_name:
                return lower_to_name[desired]
            return None
        except (OSError, ValueError, PermissionError):
            return None

    def _prompt_user_select_sheet(self, file_path: str) -> Optional[str]:
        """Prompt user to select a sheet from available options, or auto-select if only one sheet."""
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()

            if not names:
                return None

            # If there's only one sheet, automatically use it
            if len(names) == 1:
                return names[0]

            return self.ui.prompt_sheet_selection(
                file_path, names, self.processing_sheet_name
            )

        except (OSError, ValueError, PermissionError) as e:
            self.ui.show_error("Sheet Selection Error", str(e))
            return None

    # Backup helpers are provided by fileops.files; no local duplicates.
