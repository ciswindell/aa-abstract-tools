#!/usr/bin/env python3
"""
FormatExcelStep: Excel post-processing enhancement for DocumentUnit architecture.

This step applies professional formatting to Excel output files including date
formatting and auto-filter setup to improve usability and handle future data growth.

Key Features:
- Automatic date column detection and M/D/YYYY formatting
- Universal auto-filter setup with 1000-row buffer for future growth
- Graceful error handling that preserves pipeline execution
- Integration with existing DocumentUnit architecture
"""

import pandas as pd

from core.interfaces import ExcelRepo, Logger, PdfRepo, UIController
from core.pipeline.context import PipelineContext
from core.pipeline.steps import BaseStep


class FormatExcelStep(BaseStep):
    """Excel post-processing step for automatic formatting enhancement.

    This step runs after SaveStep to apply professional formatting to Excel output files.
    It automatically detects date columns and applies M/D/YYYY formatting, sets up
    auto-filters for all columns with a 1000-row buffer, and handles errors gracefully
    to ensure the main pipeline continues even if formatting fails.

    Features:
        - Case-insensitive date column detection (headers containing "Date")
        - M/D/YYYY number formatting without leading zeros
        - Auto-filter setup with 1000-row buffer for future data growth
        - Graceful error handling with warning logs
        - Works with both single-file and merge workflows

    Dependencies:
        - openpyxl library for Excel file manipulation
        - pandas DataFrame for row count calculation
        - Pipeline logging infrastructure for status messages
    """

    def __init__(
        self,
        excel_repo: ExcelRepo,
        pdf_repo: PdfRepo,
        logger: Logger,
        ui: UIController,
    ) -> None:
        """Initialize FormatExcelStep with required dependencies.

        Args:
            excel_repo: Excel repository for file operations
            pdf_repo: PDF repository (inherited requirement)
            logger: Logger for step execution messages
            ui: UI controller for user interactions
        """
        super().__init__(excel_repo, pdf_repo, logger, ui)

    def execute(self, context: PipelineContext) -> None:
        """Apply Excel formatting to saved output file.

        This method performs post-processing formatting on the Excel file saved by SaveStep.
        It applies date formatting to columns containing "Date" in the header and sets up
        auto-filters for all columns with a 1000-row buffer.

        The method implements graceful error handling - formatting failures are logged as
        warnings but do not cause the pipeline to fail, ensuring users still receive
        their core processing results.

        Args:
            context: Pipeline context containing excel_out_path (file to format) and
                    df (DataFrame for row count calculation)

        Note:
            If context.excel_out_path is None or empty, the method logs a warning and
            returns early without attempting formatting.
        """
        if not context.excel_out_path:
            return

        self.logger.info(
            f"Step {context.step_number} of {context.total_steps}: Formatting Excel..."
        )

        try:
            self._format_excel_file(context.excel_out_path, context.df)
        except Exception as e:
            # Log as error but don't fail the pipeline
            self.logger.error(f"Excel formatting failed: {e}")
            self.logger.info(
                "Pipeline will continue - core processing completed successfully"
            )

    def _find_date_columns(self, worksheet) -> list[int]:
        """Find column indices that contain 'Date' in header using robust pattern matching.

        This method performs case-insensitive pattern matching to identify date columns
        by examining the header row (row 1) for any cell containing the word "Date".
        It handles various naming patterns like "Date", "Document Date", "Received Date", etc.

        Args:
            worksheet: openpyxl worksheet object to examine

        Returns:
            List of 1-based column indices that contain date headers.
            Empty list if no date columns are found.

        Examples:
            Headers like "Date", "DATE", "Document Date", "Received Date",
            "date", "Date Received" will all be detected.
        """
        date_columns = []

        # Examine header row (row 1) for date patterns
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value:
                header_text = str(cell.value).strip().lower()
                # Case-insensitive check for "date" anywhere in the header
                if "date" in header_text:
                    date_columns.append(col_idx)

        return date_columns

    def _apply_date_formatting(
        self, worksheet, date_columns: list[int], data_rows: int
    ) -> None:
        """Apply M/D/YYYY formatting to date columns with 1000-row buffer.

        This method applies Excel number formatting to the specified date columns
        while preserving existing cell alignment and other formatting properties.
        It formats both the actual data rows and an additional 1000-row buffer to
        handle future data growth without requiring manual formatting.

        Args:
            worksheet: openpyxl worksheet object to format
            date_columns: List of 1-based column indices containing date data
            data_rows: Number of actual data rows (excluding header)

        Note:
            The M/D/YYYY format displays dates without leading zeros (e.g., 4/1/1969).
            The 1000-row buffer ensures formatting persists when new data is added.
            Existing cell alignment and other formatting properties are preserved.
        """
        if not date_columns:
            return

        # Apply formatting to data rows + 1000 buffer
        total_rows = data_rows + 1000
        formatted_cells = 0

        for col_idx in date_columns:
            for row_idx in range(2, total_rows + 2):  # Skip header row (row 1)
                cell = worksheet.cell(row=row_idx, column=col_idx)
                # Only change the number format, preserve all other formatting
                cell.number_format = "M/D/YYYY"
                formatted_cells += 1

    def _setup_auto_filter(self, worksheet, data_rows: int) -> None:
        """Setup auto-filter for all columns with 1000-row buffer.

        This method applies Excel auto-filter to all columns in the worksheet,
        extending the filter range to include actual data rows plus a 1000-row
        buffer to handle future data growth without requiring manual filter updates.

        Args:
            worksheet: openpyxl worksheet object to configure
            data_rows: Number of actual data rows (excluding header)

        Note:
            The filter range starts at A1 (header row) and extends to the last
            column and data_rows + 1000. This ensures filters work correctly
            even when new columns or rows are added to the template.
        """
        if worksheet.max_column == 0:
            return

        # Calculate filter range: A1 to last_column + buffer_rows
        total_rows = data_rows + 1000 + 1  # +1 for header row
        last_column_letter = self._get_column_letter(worksheet.max_column)
        filter_range = f"A1:{last_column_letter}{total_rows}"

        try:
            # Apply auto-filter to the calculated range
            worksheet.auto_filter.ref = filter_range
        except Exception as e:
            self.logger.error(f"Failed to apply auto-filter: {e}")
            raise

    def _get_column_letter(self, column_index: int) -> str:
        """Convert 1-based column index to Excel column letter (A, B, ..., AA, AB, ...).

        Args:
            column_index: 1-based column index

        Returns:
            Excel column letter string

        Examples:
            1 -> 'A', 26 -> 'Z', 27 -> 'AA', 52 -> 'AZ', 53 -> 'BA'
        """
        from openpyxl.utils import get_column_letter

        return get_column_letter(column_index)

    def _format_excel_file(self, excel_path: str, df: pd.DataFrame | None) -> None:
        """Apply date formatting and auto-filter to Excel file.

        This helper method performs the actual Excel formatting operations:
        1. Opens the Excel file using openpyxl
        2. Detects date columns (headers containing "Date")
        3. Applies M/D/YYYY formatting to date columns
        4. Sets up auto-filter for all columns with 1000-row buffer
        5. Saves and closes the workbook

        Args:
            excel_path: Absolute path to the Excel file to format
            df: DataFrame containing the processed data, used to determine the number
                of data rows for buffer calculation. If None, defaults to 0 rows.

        Raises:
            Exception: If Excel file cannot be opened, formatted, or saved. Exceptions
                      are caught by the calling execute() method and logged as warnings.

        Note:
            The 1000-row buffer ensures auto-filters work correctly even when new
            data is added to the template in the future.
        """
        from openpyxl import load_workbook

        # Determine data row count
        data_rows = len(df) if df is not None else 0

        # Open the Excel file
        wb = load_workbook(excel_path)
        try:
            ws = wb.active

            # 1. Find and format date columns
            date_columns = self._find_date_columns(ws)
            if date_columns:
                self._apply_date_formatting(ws, date_columns, data_rows)

            # 2. Setup auto-filter for all columns
            self._setup_auto_filter(ws, data_rows)

            # Save and close
            wb.save(excel_path)

        finally:
            wb.close()
