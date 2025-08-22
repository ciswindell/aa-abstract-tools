#!/usr/bin/env python3
"""
Excel processing module for the Abstract Renumber Tool.
Handles Excel file loading, validation, and data operations.
"""

from dataclasses import dataclass
from collections import Counter
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from dateutil.parser import parse

from excel_formatter import ExcelFormatter


@dataclass
class ColumnInfo:
    """Track all information about a column through the processing pipeline."""

    original_name: str  # Original name from source file
    current_name: str  # Current name after any mapping
    position: int  # Column position (0-based)
    width: float  # Column width from source
    horizontal_alignment: str = (
        "left"  # Horizontal alignment (left, center, right, general)
    )
    vertical_alignment: str = "top"  # Vertical alignment (top, center, bottom)
    excel_letter: str = ""  # Excel column letter (A, B, C, etc.)

    def __post_init__(self):
        """Calculate Excel letter from position."""
        if not self.excel_letter:
            self.excel_letter = self._position_to_excel_letter(self.position)

    def _position_to_excel_letter(self, pos: int) -> str:
        """Convert position to Excel column letter (0=A, 1=B, etc.)."""
        result = ""
        col = pos
        while col >= 0:
            result = chr(col % 26 + ord("A")) + result
            col = col // 26 - 1
        return result


class ExcelProcessor:
    """Handles Excel file operations and validation.

    This class processes Excel files with support for alphanumeric Index# values.
    Index# values are treated as strings throughout the processing pipeline to support
    formats like "A1", "B5", "AGH42", etc., while maintaining backward compatibility
    with numeric-only Index# values.
    """

    def __init__(self, required_columns: List[str]):
        self.required_columns = required_columns
        self.df: Optional[pd.DataFrame] = None
        self.original_file_path: Optional[str] = None
        self.processed_columns = set()  # Track which columns we've modified
        self.bookmark_formula_column: Optional[str] = None
        self.columns: List[ColumnInfo] = []  # Complete column information
        self.source_column_widths_by_position = (
            {}
        )  # Store original column widths by position

    def load_file(self, file_path: str, sheet_name: Optional[str] = None) -> bool:
        """Load Excel file with basic validation.

        Args:
            file_path: Path to the Excel file
            sheet_name: Optional sheet name to load (case-sensitive exact as in file)
        """
        try:
            # Load Excel file
            self.df = pd.read_excel(file_path, dtype=str, sheet_name=sheet_name)
            self.original_file_path = file_path

            if self.df.empty:
                raise ValueError("Excel file contains no data")

            # Initialize column information
            self._initialize_column_info()

            # Add Original_Index column
            self._add_original_index_column()

            # Detect bookmark formula column
            self._detect_bookmark_formula_column()

            # Process data types
            self._process_data_types()

            return True

        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {str(e)}")

    def _initialize_column_info(self) -> None:
        """Initialize column information with source formatting from Excel file."""
        self.columns = []

        # Extract column widths and alignments from source Excel file
        source_widths = self._extract_source_column_widths()
        source_alignments = self._extract_source_alignments()

        # Create ColumnInfo for each column
        for pos, col_name in enumerate(self.df.columns):
            width = source_widths.get(pos, 15.0)  # Default width if not found
            h_align, v_align = source_alignments.get(pos, ("left", "top"))

            col_info = ColumnInfo(
                original_name=col_name,
                current_name=col_name,
                position=pos,
                width=width,
                horizontal_alignment=h_align,
                vertical_alignment=v_align,
                excel_letter="",  # Will be calculated in __post_init__
            )
            self.columns.append(col_info)

    def _extract_source_column_widths(self) -> Dict[int, float]:
        """Extract column widths from source Excel file."""
        widths = {}

        if not self.original_file_path:
            return widths

        try:
            wb = load_workbook(self.original_file_path)
            ws = wb.active

            for pos in range(len(self.df.columns)):
                col_letter = self._get_excel_column_letter(pos)
                width = ws.column_dimensions[col_letter].width
                widths[pos] = width if width else 15.0

            wb.close()
        except:
            # Use defaults if extraction fails
            for pos in range(len(self.df.columns)):
                widths[pos] = 15.0

        return widths

    def _extract_source_alignments(self) -> Dict[int, tuple]:
        """Extract column alignments from source Excel file."""
        alignments = {}

        if not self.original_file_path:
            return alignments

        try:
            wb = load_workbook(self.original_file_path)
            ws = wb.active

            for pos in range(len(self.df.columns)):
                col_letter = self._get_excel_column_letter(pos)
                cell = ws[f"{col_letter}2"]  # Check first data row

                h_align = "left"
                v_align = "top"

                if cell.alignment:
                    if cell.alignment.horizontal:
                        h_align = str(cell.alignment.horizontal)
                    if cell.alignment.vertical:
                        v_align = str(cell.alignment.vertical)

                alignments[pos] = (h_align, v_align)

            wb.close()
        except:
            # Use defaults if extraction fails
            for pos in range(len(self.df.columns)):
                alignments[pos] = ("left", "top")

        return alignments

    def _detect_bookmark_formula_column(self) -> None:
        """Detect which column contains bookmark formulas."""
        if self.df is None:
            return

        # Look for a column that might contain bookmark formulas
        # Common names: "Bookmark", "Bookmark Formula", "Formula", etc.
        possible_names = ["Bookmark", "Bookmark Formula", "Formula", "Bookmark Text"]

        for col in self.df.columns:
            if any(name.lower() in col.lower() for name in possible_names):
                self.bookmark_formula_column = col
                break

    def _get_excel_column_letter(self, col_index: int) -> str:
        """Convert column index to Excel column letter (0=A, 1=B, etc.)."""
        result = ""
        while col_index >= 0:
            result = chr(col_index % 26 + ord("A")) + result
            col_index = col_index // 26 - 1
        return result

    def _process_data_types(self) -> None:
        """Process data types for specific columns to optimize sorting and display.

        Index# column is treated as string type to support alphanumeric values.
        Original Index# values are preserved for PDF bookmark mapping.
        """
        if self.df is None:
            return

        # Handle Index# column - keep as string to support alphanumeric values like A1, B1, etc.
        if "Index#" in self.df.columns:
            # Ensure Index# is treated as string to handle alphanumeric values
            self.df["Index#"] = self.df["Index#"].astype(str).str.strip()
            self.df["Index#"] = self.df["Index#"].replace("nan", "")

            # Additional cleaning and validation for Index# values
            self.df["Index#"] = self.df["Index#"].str.replace(
                r"\s+", "", regex=True
            )  # Remove all whitespace
            self.df["Index#"] = self.df["Index#"].replace(
                "", "N/A"
            )  # Replace empty strings with placeholder

            # Check for duplicate Index# values (for validation only - don't modify to preserve PDF bookmark matching)
            if self.df["Index#"].duplicated().any():
                # Log warning about duplicates but don't modify values to preserve PDF bookmark matching
                duplicate_values = self.df["Index#"][
                    self.df["Index#"].duplicated()
                ].unique()
                print(
                    f"Warning: Duplicate Index# values found: {duplicate_values}. These may cause issues with PDF bookmark mapping."
                )

            self.processed_columns.add("Index#")

        # Process date columns with robust handling to prevent data loss
        date_columns = ["Document Date", "Received Date"]
        for col in date_columns:
            if col in self.df.columns:
                self._process_date_column_robust(col)
                self.processed_columns.add(col)

        # Clean text columns
        text_columns = ["Legal Description", "Grantee", "Grantor", "Document Type"]
        for col in text_columns:
            if col in self.df.columns:
                try:
                    # Ensure string type and strip whitespace
                    self.df[col] = self.df[col].astype(str).str.strip()
                    # Handle NaN values
                    self.df[col] = self.df[col].replace("nan", "")
                    self.processed_columns.add(col)
                except (ValueError, TypeError, AttributeError):
                    pass

    def _process_date_column_robust(self, col: str) -> None:
        """
        Robust date processing that prevents data loss.

        Strategy:
        1. Try multiple parsing approaches
        2. Use original value if all parsing fails
        3. Never lose data
        """
        if col not in self.df.columns:
            return

        # Process each value individually for maximum robustness
        processed_values = []

        for idx, value in enumerate(self.df[col]):
            processed_value = self._parse_date_value_robust(value, idx, col)
            processed_values.append(processed_value)

        # Update the column with processed values
        self.df[col] = processed_values

        # Log conversion summary
        converted_count = sum(
            1 for v in processed_values if isinstance(v, (pd.Timestamp, datetime))
        )
        original_count = len(processed_values)
        non_null_count = sum(
            1 for v in processed_values if pd.notna(v) and str(v).strip() != ""
        )

        # Date processing completed silently

    def _parse_date_value_robust(self, value, idx: int, col_name: str):
        """
        Robust date parsing that tries multiple approaches and never loses data.

        Args:
            value: The original date value
            idx: Row index for debugging
            col_name: Column name for debugging

        Returns:
            Parsed datetime or original value if parsing fails
        """
        # Handle null/empty values
        if pd.isna(value) or value is None:
            return value

        # If already a datetime object, return as-is
        if isinstance(value, (datetime, pd.Timestamp)):
            return value

        # Convert to string for processing
        str_value = str(value).strip()

        # Handle empty strings
        if str_value == "" or str_value.lower() == "nan":
            return value

        # Try pandas default parsing first (handles most formats)
        try:
            parsed = pd.to_datetime(str_value, errors="raise")
            if pd.notna(parsed):
                return parsed
        except:
            pass

        # Try common date formats manually
        common_formats = [
            "%m/%d/%Y",  # 1/30/1959
            "%m/%d/%y",  # 1/30/59
            "%m-%d-%Y",  # 1-30-1959
            "%m-%d-%y",  # 1-30-59
            "%Y-%m-%d",  # 1959-01-30
            "%Y/%m/%d",  # 1959/01/30
            "%d/%m/%Y",  # 30/1/1959
            "%d-%m-%Y",  # 30-1-1959
            "%B %d, %Y",  # January 30, 1959
            "%b %d, %Y",  # Jan 30, 1959
            "%Y-%m-%d %H:%M:%S",  # 1959-01-30 00:00:00
        ]

        for fmt in common_formats:
            try:
                parsed = datetime.strptime(str_value, fmt)
                return pd.Timestamp(parsed)
            except:
                continue

        # Try with dateutil parser (very flexible)
        try:
            parsed = parse(str_value)
            return pd.Timestamp(parsed)
        except:
            pass

        # If all parsing fails, keep original value silently
        return value

    def check_duplicate_columns(self) -> List[str]:
        """Check for duplicate column names in the loaded DataFrame."""
        return self.df.columns[self.df.columns.duplicated()].tolist()

    def get_missing_columns(self) -> List[str]:
        """Check for required columns and return list of missing ones."""
        if self.df is None:
            return self.required_columns
        present_lower = {str(c).lower() for c in self.df.columns}
        missing = [
            col
            for col in self.required_columns
            if str(col).lower() not in present_lower
        ]
        return missing

    def get_required_duplicate_columns(self) -> List[str]:
        """Return required column names that appear more than once (case-insensitive)."""
        if self.df is None:
            return []
        required_lower = {str(c).lower() for c in self.required_columns}
        counts = Counter(
            str(c).lower() for c in self.df.columns if str(c).lower() in required_lower
        )
        return [
            rc for rc in self.required_columns if counts.get(str(rc).lower(), 0) > 1
        ]

    def apply_column_mapping(self, mapping: Dict[str, str]) -> bool:
        """Apply column name mappings to the DataFrame."""
        # Update column info with new mappings
        for col_info in self.columns:
            if col_info.current_name in mapping:
                col_info.current_name = mapping[col_info.current_name]

        # Apply mappings to DataFrame
        self.df.rename(columns=mapping, inplace=True)

        # Verify all required columns are now present
        still_missing = self.get_missing_columns()
        if still_missing:
            raise ValueError(f"Still missing columns after mapping: {still_missing}")

        # Reprocess data types after column mapping
        self._process_data_types()
        return True

    def get_column_info(self) -> Dict[str, any]:
        """Get information about the loaded DataFrame."""
        if self.df is None:
            return {}

        return {
            "rows": len(self.df),
            "columns": len(self.df.columns),
            "column_names": self.df.columns.tolist(),
        }

    def get_dataframe(self) -> Optional[pd.DataFrame]:
        """Get the loaded DataFrame.

        Returns:
            Optional[pd.DataFrame]: The loaded pandas DataFrame, or None if no file loaded
        """
        return self.df

    def sort_data(self) -> bool:
        """Sort the DataFrame according to PRD requirements."""
        # Sort order: Received Date, Document Date, Document Type, Grantor, Grantee, Legal Description
        sort_columns = [
            "Received Date",
            "Document Date",
            "Document Type",
            "Grantor",
            "Grantee",
            "Legal Description",
        ]

        # Filter to only include columns that exist in the DataFrame
        existing_sort_columns = [col for col in sort_columns if col in self.df.columns]

        # Handle null/missing values before sorting
        self._handle_missing_values()

        # Ensure proper data types for sorting columns before sorting
        self._prepare_sorting_data_types()

        # Sort by multiple columns - all ascending
        self.df = self.df.sort_values(
            by=existing_sort_columns, ascending=True, na_position="last"
        )
        # Reset index to maintain proper row order
        self.df = self.df.reset_index(drop=True)

        # Renumber Index# column starting from 1
        self._renumber_index()
        return True

    def _prepare_sorting_data_types(self) -> None:
        """Ensure proper data types for sorting columns."""
        # Clean and prepare text columns for proper sorting
        text_columns = ["Legal Description", "Grantee", "Grantor", "Document Type"]
        for col in text_columns:
            if col in self.df.columns:
                self.df[col] = self.df[col].astype(str).str.strip()
                self.df[col] = self.df[col].replace("nan", "")

        # Ensure date columns are properly formatted for sorting
        # Note: Date columns are already processed robustly in _process_data_types()
        # This just ensures any remaining issues are handled
        date_columns = ["Document Date", "Received Date"]
        for col in date_columns:
            if col in self.df.columns:
                # Check if any values are still not datetime/timestamp
                non_datetime_mask = ~self.df[col].apply(
                    lambda x: isinstance(x, (pd.Timestamp, datetime)) or pd.isna(x)
                )
                if non_datetime_mask.any():
                    # Re-process date column silently
                    self._process_date_column_robust(col)

    def _renumber_index(self) -> None:
        """Renumber the Index# column starting from 1 and incrementing by 1.

        This method always creates sequential integer values (1, 2, 3, etc.) regardless
        of the original Index# format (numeric, alphanumeric, or mixed).
        """
        if "Index#" in self.df.columns:
            # Always create sequential integers for the final Index# column
            self.df["Index#"] = range(1, len(self.df) + 1)
            self.processed_columns.add("Index#")

    def _handle_missing_values(self) -> None:
        """Handle null/missing values in sorting columns."""
        text_columns = ["Legal Description", "Grantee", "Grantor", "Document Type"]
        for col in text_columns:
            if col in self.df.columns:
                self.df[col] = self.df[col].fillna("")
                self.processed_columns.add(col)

    def get_processed_columns(self) -> set:
        """Return set of columns that have been processed/modified.

        Returns:
            set: Set of column names that have been processed or modified
        """
        return self.processed_columns.copy()

    def get_unprocessed_columns(self) -> List[str]:
        """Return list of columns that haven't been processed.

        Returns:
            List[str]: List of column names that have not been processed or modified
        """
        if self.df is None:
            return []
        return [col for col in self.df.columns if col not in self.processed_columns]

    def save_with_formulas(
        self, output_path: str, processing_sheet_name: Optional[str] = None
    ) -> bool:
        """Save Excel file preserving formatting.

        Writes values into the existing output workbook (a copy of the input)
        to retain all sheets and formatting. Only the active/processing sheet is modified.
        """
        try:
            # Create output DataFrame with original column names
            output_df = self._create_output_dataframe()

            # Open the copied workbook and select the processing sheet or fallback to active
            wb = load_workbook(output_path)
            # Require explicit processing sheet; no fallback to active sheet
            if not processing_sheet_name:
                wb.close()
                raise RuntimeError("Processing sheet name not provided")
            lower_to_name = {name.lower(): name for name in wb.sheetnames}
            match_name = lower_to_name.get(str(processing_sheet_name).lower())
            if not match_name:
                wb.close()
                raise RuntimeError(
                    f"Processing sheet '{processing_sheet_name}' not found in output workbook"
                )
            ws = wb[match_name]

            # Build a case-insensitive header map from the first row
            header_values = []
            for cell in ws[1]:
                header_values.append(
                    "" if cell.value is None else str(cell.value).strip()
                )
            header_to_col = {
                h.lower(): idx + 1 for idx, h in enumerate(header_values) if h != ""
            }

            # Write DataFrame values into matching columns by header name
            for df_col in output_df.columns:
                target_col_idx = header_to_col.get(str(df_col).strip().lower())
                if not target_col_idx:
                    # Skip columns that do not exist in the target sheet
                    continue

                values = output_df[df_col].tolist()
                for row_idx, value in enumerate(
                    values, start=2
                ):  # Start at row 2 (after header)
                    ws.cell(row=row_idx, column=target_col_idx, value=value)

            # Clear hard-coded fills (keep conditional formatting)
            empty_fill = PatternFill(fill_type=None)
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, max_col=ws.max_column
            ):
                for cell in row:
                    cell.fill = empty_fill

            # Keep the selected sheet as active for downstream formatting
            try:
                wb.active = wb.sheetnames.index(ws.title)
            except Exception:
                pass

            # Save workbook with updated values
            wb.save(output_path)
            wb.close()

            # Update column positions for the final output
            self._update_column_positions_for_output(output_df)

            # Apply formatting (only affects the active/processing sheet)
            formatter = ExcelFormatter(self.columns, self.bookmark_formula_column)
            formatter.apply_formatting(output_path, output_df)
            return True

        except Exception as e:
            raise RuntimeError(f"Failed to save Excel file: {str(e)}") from e

    def _create_output_dataframe(self) -> pd.DataFrame:
        """Create DataFrame with original column names for export, excluding internal columns."""
        output_df = self.df.copy()

        # Remove the Original_Index column - it's only used internally for PDF bookmark mapping
        # and should not appear in the final Excel output
        if "Original_Index" in output_df.columns:
            output_df = output_df.drop(columns=["Original_Index"])

        # Create mapping from current names back to original names
        name_mapping = {}
        for col_info in self.columns:
            # Only include columns that are not internal/temporary columns
            if (
                col_info.current_name in output_df.columns
                and col_info.original_name != "Original_Index"
            ):
                name_mapping[col_info.current_name] = col_info.original_name

        # Rename columns back to original names
        output_df.rename(columns=name_mapping, inplace=True)

        return output_df

    def _update_column_positions_for_output(self, output_df: pd.DataFrame) -> None:
        """Update column positions to match the output DataFrame column order."""
        # Create a mapping from column names to their new positions in the output DataFrame
        position_mapping = {name: pos for pos, name in enumerate(output_df.columns)}

        # Update each column's position
        for col_info in self.columns:
            # Skip internal columns that don't appear in final output
            if col_info.original_name == "Original_Index":
                continue

            # Only update columns that exist in the output DataFrame
            if col_info.current_name in position_mapping:
                col_info.position = position_mapping[col_info.current_name]

    def _find_column_by_current_name(self, current_name: str) -> Optional[ColumnInfo]:
        """Find column info by current (mapped) name."""
        for col_info in self.columns:
            if col_info.current_name == current_name:
                return col_info
        return None

    def get_bookmark_formula_column(self) -> Optional[str]:
        """Return the name of the bookmark formula column if detected.

        Returns:
            Optional[str]: Name of the detected bookmark formula column, or None if not found
        """
        return self.bookmark_formula_column

    def _add_original_index_column(self) -> None:
        """Add Original_Index column to preserve original Index# values for PDF bookmark mapping."""
        if "Index#" in self.df.columns:
            # Ensure Index# values are preserved as strings for alphanumeric support
            self.df["Original_Index"] = self.df["Index#"].astype(str).copy()
            self.processed_columns.add("Original_Index")

            # Add column info for Original_Index
            original_index_col_info = ColumnInfo(
                original_name="Original_Index",
                current_name="Original_Index",
                position=len(self.columns),
                width=0,
                horizontal_alignment="left",
                vertical_alignment="top",
                excel_letter="",
            )
            self.columns.append(original_index_col_info)

    def get_original_index_mapping(self) -> Dict[str, int]:
        """Get mapping from original index values to new index values.

        Returns:
            Dict[str, int]: Mapping from original index (as string) to new index (as int)
        """
        if (
            self.df is None
            or "Original_Index" not in self.df.columns
            or "Index#" not in self.df.columns
        ):
            return {}

        # Create mapping: original_index (as string) -> new_index (as int)
        mapping = {}
        for _, row in self.df.iterrows():
            original_idx = str(row["Original_Index"]).strip()
            new_idx = int(row["Index#"])
            mapping[original_idx] = new_idx

        return mapping
