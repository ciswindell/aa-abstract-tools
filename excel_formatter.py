#!/usr/bin/env python3
"""
Excel formatting module for the Abstract Renumber Tool.
Handles all Excel file formatting operations including alignment, widths, and styling.
"""

from typing import Any, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils.excel_utils import index_to_col_letter


class ExcelFormatter:
    """Handles Excel file formatting operations."""

    def __init__(
        self, columns: List[Any], bookmark_formula_column: Optional[str] = None
    ) -> None:
        self.columns = columns
        self.bookmark_formula_column = bookmark_formula_column

    def apply_formatting(self, workbook_path: str, output_df: pd.DataFrame) -> bool:
        """Apply minimal updates: formulas only; template formatting is authoritative."""
        try:
            wb = load_workbook(workbook_path)
            ws = wb.active

            # Apply bookmark formulas if needed
            if self.bookmark_formula_column:
                self._apply_bookmark_formulas(ws, output_df)

            # Force Excel to recalculate formulas on open
            try:
                wb.calculation_properties.fullCalcOnLoad = True  # type: ignore[attr-defined]
            except Exception:
                pass

            wb.save(workbook_path)
            wb.close()

            return True

        except Exception as e:
            raise ValueError(f"Failed to apply formatting: {str(e)}")

    def apply_formulas_only(self, workbook_path: str, output_df: pd.DataFrame) -> bool:
        """Apply only bookmark formulas; leave existing formatting intact."""
        try:
            wb = load_workbook(workbook_path)
            ws = wb.active

            if self.bookmark_formula_column:
                self._apply_bookmark_formulas(ws, output_df)

            try:
                wb.calculation_properties.fullCalcOnLoad = True  # type: ignore[attr-defined]
            except Exception:
                pass

            wb.save(workbook_path)
            wb.close()
            return True
        except Exception as e:
            raise ValueError(f"Failed to apply formulas: {str(e)}")

    # Removed: explicit width/align/wrap/filters/freeze; rely on template formatting

    def _get_column_letter_by_position(self, position: int) -> str:
        """Convert position to Excel column letter (0=A, 1=B, etc.)."""
        return index_to_col_letter(position)

    # Removed: cell alignment, wrapping, date formats, auto-filters, freeze panes

    def _apply_bookmark_formulas(
        self, worksheet: Worksheet, output_df: pd.DataFrame
    ) -> None:
        """Apply bookmark formulas using original column structure."""
        try:
            # Find bookmark column info
            bookmark_col_info = None
            for col_info in self.columns:
                if col_info.current_name == self.bookmark_formula_column:
                    bookmark_col_info = col_info
                    break

            if not bookmark_col_info:
                return

            # Find required column positions in original structure
            index_col_info = self._find_column_by_current_name("Index#")
            doc_type_col_info = self._find_column_by_current_name("Document Type")
            received_date_col_info = self._find_column_by_current_name("Received Date")

            if all([index_col_info, doc_type_col_info, received_date_col_info]):
                # Generate formulas using original column letters
                for row_num in range(2, len(output_df) + 2):
                    formula = (
                        f'={self._get_column_letter_by_position(index_col_info.position)}{row_num}&"-"&'
                        f'{self._get_column_letter_by_position(doc_type_col_info.position)}{row_num}&"-"&'
                        f'TEXT({self._get_column_letter_by_position(received_date_col_info.position)}{row_num},"m/d/yyyy")'
                    )

                    cell = f"{self._get_column_letter_by_position(bookmark_col_info.position)}{row_num}"
                    worksheet[cell] = formula

        except Exception:
            pass  # Skip if bookmark formulas can't be applied

    def _find_column_by_current_name(self, current_name: str) -> Optional[Any]:
        """Find column info by current (mapped) name."""
        for col_info in self.columns:
            if col_info.current_name == current_name:
                return col_info
        return None
