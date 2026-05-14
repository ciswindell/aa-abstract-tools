#!/usr/bin/env python3
"""
Excel repository adapter using pandas and openpyxl.

Loads a worksheet into a DataFrame and saves DataFrame values
back into an existing workbook template and sheet, then applies
formatting via `ExcelFormatter`.
"""

from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from fileops.files import atomic_write_with_template
from utils.bookmark_formulas import (
    apply_bookmark_formulas,
    detect_bookmark_column,
    has_bookmark_formulas,
)
from utils.dates import parse_robust


class ExcelOpenpyxlRepo:
    """Concrete ExcelRepo implementation using pandas/openpyxl."""

    def load(self, path: str, sheet: Optional[str]) -> pd.DataFrame:
        """Load a worksheet into a DataFrame preserving native Excel data types.

        Date columns remain as datetime objects, while Index# is converted to string
        for consistent bookmark matching. Removes completely empty rows to prevent
        processing errors.
        """
        df = pd.read_excel(path, sheet_name=sheet)

        # Remove completely empty rows (all columns are NaN/empty)
        # This prevents processing errors while preserving rows with any data
        if not df.empty:
            df = df.dropna(how="all").reset_index(drop=True)

        # Convert Index# column to string for consistent bookmark matching
        if "Index#" in df.columns:
            df["Index#"] = (
                df["Index#"].fillna("").astype(str).str.strip().replace("nan", "")
            )

        # Convert date columns that may be formatted as text in Excel
        # This ensures chronological sorting works correctly
        date_column_patterns = ["date"]
        for col in df.columns:
            col_lower = str(col).lower()
            if any(pattern in col_lower for pattern in date_column_patterns):
                df[col] = df[col].apply(parse_robust)

        return df

    def get_sheet_names(self, path: str) -> list[str]:
        """Get list of sheet names in the Excel file."""
        wb = load_workbook(path, read_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()

    def _normalize_column_name(self, name: str) -> str:
        """
        Normalize column name for matching.

        Converts to lowercase and normalizes whitespace (trim and collapse).

        Args:
            name: Raw column name

        Returns:
            Normalized name for case-insensitive, whitespace-tolerant matching

        Example:
            " Status  " -> "status"
            "Date Created" -> "date created"
        """
        return " ".join(str(name).strip().lower().split())

    def _write_dataframe_to_workbook(
        self,
        temp_path: str,
        df: pd.DataFrame,
        target_sheet: str,
        bookmark_col_name: Optional[str],
        add_missing_columns: bool = False,
    ) -> None:
        """
        Write DataFrame to workbook and apply bookmark formulas if needed.

        Supports dynamic column addition for merge workflows. When add_missing_columns
        is True, new columns from the DataFrame that don't exist in the template are
        added to the output, excluding system columns (those starting with underscore
        or named Document_ID).

        Column matching is case-insensitive and whitespace-tolerant, allowing proper
        merging of files with minor column name variations.

        Args:
            temp_path: Path to temporary workbook file
            df: DataFrame to write
            target_sheet: Name of sheet to write to
            bookmark_col_name: Optional bookmark column to skip
            add_missing_columns: If True, adds new columns from df (excluding system columns)
        """
        wb = load_workbook(temp_path)
        try:
            # Find target sheet
            lower_to_name = {name.lower(): name for name in wb.sheetnames}
            sheet_name = lower_to_name.get(str(target_sheet).lower())
            if not sheet_name:
                raise RuntimeError(
                    f"Processing sheet '{target_sheet}' not found in output workbook"
                )
            ws = wb[sheet_name]

            # Build header map with normalized column names
            header_values = [
                "" if cell.value is None else str(cell.value).strip() for cell in ws[1]
            ]
            header_to_col = {
                self._normalize_column_name(h): idx + 1
                for idx, h in enumerate(header_values)
                if h != ""
            }

            # Handle new columns that should be added if missing
            if add_missing_columns:
                # Define system columns that should never be added to output
                # These are internal tracking columns not meant for user visibility
                SYSTEM_COLUMNS = {"_include", "_original_index", "Document_ID"}

                # Find all new columns that don't exist in template (excluding system columns)
                new_columns = [
                    col
                    for col in df.columns
                    if self._normalize_column_name(col) not in header_to_col
                    and col not in SYSTEM_COLUMNS
                ]

                # Add new columns to the Excel template header row
                if new_columns:
                    next_col_idx = (
                        max(header_to_col.values()) + 1 if header_to_col else 1
                    )
                    for new_col in new_columns:
                        ws.cell(row=1, column=next_col_idx, value=new_col)
                        header_to_col[self._normalize_column_name(new_col)] = (
                            next_col_idx
                        )
                        next_col_idx += 1

                    # Note: Logging would require logger dependency - skip for now to keep class simple

            # Bound clearing loops by the real data extent, not ws.max_row.
            # Some templates report phantom dimensions (max_row near Excel's
            # 1M-row limit) from a stray cell touched far below the data. The
            # max() across populated cells in our header columns gives the true
            # extent in microseconds; combined with len(df)+1 it covers both
            # the filter case (df smaller than template) and the merge case
            # (df larger than template).
            data_col_indices = set(header_to_col.values())
            real_max_row = max(
                (
                    r
                    for (r, c), cell in ws._cells.items()
                    if c in data_col_indices and cell.value is not None
                ),
                default=1,
            )
            clear_max_row = max(real_max_row, len(df) + 1)

            # Clear existing data cell values (preserve formatting)
            if clear_max_row > 1:
                for row_idx in range(2, clear_max_row + 1):
                    for col_idx in header_to_col.values():
                        ws.cell(row=row_idx, column=col_idx, value="")

            # Write DataFrame values
            for df_col in df.columns:
                if bookmark_col_name and self._normalize_column_name(
                    df_col
                ) == self._normalize_column_name(bookmark_col_name):
                    continue  # Skip bookmark formula column
                col_idx = header_to_col.get(self._normalize_column_name(df_col))
                if col_idx:
                    # Convert boolean values to "Yes"/"No" for Document_Found column
                    if df_col == "Document_Found":
                        for row_idx, value in enumerate(df[df_col].tolist(), start=2):
                            excel_value = "Yes" if value else "No"
                            ws.cell(row=row_idx, column=col_idx, value=excel_value)
                    else:
                        for row_idx, value in enumerate(df[df_col].tolist(), start=2):
                            ws.cell(row=row_idx, column=col_idx, value=value)

            # Clear fills across the same row range and the header columns.
            empty_fill = PatternFill(fill_type=None)
            last_col = max(header_to_col.values()) if header_to_col else 1
            for row in ws.iter_rows(
                min_row=1, max_row=clear_max_row, max_col=last_col
            ):
                for cell in row:
                    cell.fill = empty_fill
            wb.save(temp_path)
        finally:
            wb.close()

        # Apply bookmark formulas if needed
        if bookmark_col_name and not has_bookmark_formulas(
            temp_path, bookmark_col_name
        ):
            apply_bookmark_formulas(temp_path, df, bookmark_col_name)

    def save(
        self,
        df: pd.DataFrame,
        template_path: str,
        target_sheet: str,
        out_path: str,
        add_missing_columns: bool = False,
    ) -> None:
        """Atomically write DataFrame values into the workbook template and sheet.

        Preserves any bookmark formula column by not overwriting its cells and
        delegating formula application to ExcelFormatter.

        Args:
            df: DataFrame to save
            template_path: Path to Excel template file
            target_sheet: Name of target sheet
            out_path: Output file path
            add_missing_columns: Whether to add whitelisted missing columns to Excel template
        """

        if not target_sheet:
            raise RuntimeError("Target sheet name not provided")

        # Detect bookmark formula column name in the DataFrame
        bookmark_col_name = detect_bookmark_column(df)

        def _write_into(temp_path: str) -> None:
            self._write_dataframe_to_workbook(
                temp_path, df, target_sheet, bookmark_col_name, add_missing_columns
            )

        atomic_write_with_template(
            template_path=template_path,
            out_path=out_path,
            write_into_path=_write_into,
        )
